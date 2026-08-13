"""Check every derived series against the workbook it replaces.

    python reconcile_test.py

This is the test that matters most. Six years of readings were kept in a
spreadsheet, and the only reason to trust the dashboard's weekly averages and
month-on-month changes is that they reproduce the ones already there. Every
figure is recomputed from the database and compared with the cached value in
the corresponding sheet.

The workbook is frozen
----------------------
It stopped being written to on 08/08/2026, when the dashboard took over, and
the database has moved on past it. That is the point of the exercise, but it
does mean the two no longer end on the same day, and every period the sheet
only partly covers now legitimately disagrees:

  * the **week** the sheet ends in - w/c 03/08/2026 - is a part-week there and
    a whole week here;
  * the **month** it ends in - August 2026 - is 8 days there and however many
    have happened here;
  * and the month-on-month change into that month, which is built on it.

These are derived from the workbook's own last reading rather than hard-coded,
so they follow it if it is ever brought up to date, and they disappear entirely
if it catches the database up. Periods the sheet does not reach at all are not
compared: there is nothing to disagree with.

Four known deviations
---------------------
They are the workbook's, not the dashboard's, and each is asserted rather than
tolerated - if one ever stops being true the test fails and says so.

1. **The Summary sheet silently drops three days.** Its column A is
   `=Data!A3` filled down, but at row 818 the references slipped by one and
   never recovered - twice more after that. The result is that 09/08/2022,
   29/01/2023 and 04/09/2025 were recorded on the Data sheet and then left out
   of every average built on Summary: the monthly figures, both weekly-change
   sheets and the monthly change sheet. The dashboard includes them, so those
   periods legitimately disagree.

2. **The rolling weekly change is a row offset, not a date offset.** The sheet
   computes `Summary!B10 - Summary!B3` - seven *rows* apart, which is seven
   days only while Summary has one row per day. Past the first dropped day it
   is comparing eight calendar days. The dashboard joins on the date.

3. **The current part-week average is divided by seven regardless.** Weekly
   Results averages a fixed seven-cell window, and the cells below the last
   reading hold formulas evaluating to 0 rather than being empty, so AVERAGE
   counts them. The part-week at the end of the data therefore reads about
   6/7 of the truth. The dashboard averages the days that exist.

4. **Monthly rows exist for months that have not happened.** Harmless - they
   are blank - but they are skipped rather than reported as missing.
"""
from __future__ import annotations

import datetime as dt
import sys
from pathlib import Path

import config
from core import queries

WORKBOOK = config.SOURCE_XLSX

# Column order on every derived sheet: date, then these six.
KEYS = ["weight", "bmi", "body_fat_pct", "skeletal_muscle_pct", "rm_kcal",
        "visceral_fat"]

# The derived sheets store the two percentages as fractions with a percent
# format, so 14.3% is held as 0.143. The Data sheet does not.
SCALE = {"body_fat_pct": 100.0, "skeletal_muscle_pct": 100.0}

# Floating point only - anything larger is a real disagreement.
TOL = 1e-5

# Deviation 1: days the Summary sheet lost.
DROPPED = [dt.date(2022, 8, 9), dt.date(2023, 1, 29), dt.date(2025, 9, 4)]

failures: list = []


def check(label: str, actual, expected) -> None:
    ok = actual == expected
    print(f"  [{'ok ' if ok else 'FAIL'}] {label}: {actual!r}"
          + ("" if ok else f" (expected {expected!r})"))
    if not ok:
        failures.append(label)


def as_date(value):
    if isinstance(value, dt.datetime):
        return value.date()
    return value if isinstance(value, dt.date) else None


def sheet_rows(wb, name: str, columns: int = 7) -> list:
    """Every row of a derived sheet that carries a real date in column A."""
    out = []
    for row in wb[name].iter_rows(min_row=3, max_col=columns, values_only=True):
        when = as_date(row[0])
        if when is not None:
            out.append((when, row))
    return out


def values_match(mine: dict, row) -> list:
    """The metrics where our figure and the sheet's disagree."""
    wrong = []
    for index, key in enumerate(KEYS, start=1):
        theirs = row[index]
        if theirs in (None, ""):
            continue
        theirs = float(theirs) * SCALE.get(key, 1.0)
        if abs(float(mine[key]) - theirs) > TOL:
            wrong.append(f"{key}: {mine[key]:.6g} vs {theirs:.6g}")
    return wrong


def compare(wb, label: str, sheet: str, ours: list, expect_absent=(),
            expect_differs=()) -> None:
    """Compare a series with its sheet, allowing for the known deviations."""
    mine = {r["period"]: r for r in ours}
    absent, differs, checked = [], [], 0

    for when, row in sheet_rows(wb, sheet):
        row_key = when.isoformat()
        if row_key not in mine:
            # A future month the sheet has pre-built, or a period we do not
            # cover. Blank rows are not a disagreement.
            if any(cell not in (None, "") for cell in row[1:]):
                differs.append(f"{row_key}: sheet has data we do not")
            continue
        checked += 1
        wrong = values_match(mine[row_key], row)
        if wrong:
            differs.append(row_key)

    sheet_days = {when for when, _ in sheet_rows(wb, sheet)}
    absent = sorted(d for d in expect_absent if d not in sheet_days)

    print(f"\n{label} ({sheet})")
    check(f"{checked} periods compared", checked > 0, True)
    check("days the sheet is missing", absent, sorted(expect_absent))
    check("periods that differ", sorted(differs), sorted(expect_differs))


# --------------------------------------------------------------------------- #
def affected_months(sheet_last: dt.date | None = None) -> list:
    """Months the sheet and the database cannot agree on.

    The months containing a day Summary lost, plus - once the database has run
    past the workbook - the month the workbook stops in, which is a part-month
    there and a fuller one here.
    """
    out = {d.replace(day=1).isoformat() for d in DROPPED}
    out |= set(trailing_month(sheet_last))
    return sorted(out)


def trailing_month(sheet_last: dt.date | None) -> list:
    """The month the workbook stops in, if the database has more of it."""
    if sheet_last is None:
        return []
    last_db = dt.date.fromisoformat(queries.coverage()["last_day"])
    if last_db <= sheet_last:
        return []
    return [sheet_last.replace(day=1).isoformat()]


def affected_month_changes(sheet_last: dt.date | None = None) -> list:
    """A month-on-month change is wrong if either month is.

    Limited to months the database actually has a change for: the month after
    the workbook's last is usually still in the future, and a change that does
    not exist cannot disagree.
    """
    have = {row["period"] for row in queries.changes("monthly")}
    out = set()
    for iso in affected_months(sheet_last):
        when = dt.date.fromisoformat(iso)
        out.add(iso)
        following = (when.replace(day=28) + dt.timedelta(days=4)).replace(day=1)
        out.add(following.isoformat())
    return sorted(out & have)


def affected_rolling() -> list:
    """The sheet's seven-row window is wrong wherever it spans a lost day.

    The lost day itself is not in the list: the sheet has no row for it at all,
    so there is nothing to disagree with. It is the seven days after it whose
    window has silently become eight days wide.
    """
    out = set()
    for dropped in DROPPED:
        for offset in range(1, 8):
            out.add((dropped + dt.timedelta(days=offset)).isoformat())
    covered = {r["period"] for r in queries.rolling_change(7)}
    return sorted(out & covered)


def affected_weeks() -> list:
    """Weeks whose seven-row window is not seven days.

    Two ways that happens: the week holds a day whose rolling change is wrong,
    or the week holds a lost day, in which case the window runs a day past the
    end of the week to make up the seven rows.
    """
    out = set()
    for when in [dt.date.fromisoformat(iso) for iso in affected_rolling()] + DROPPED:
        out.add((when - dt.timedelta(days=when.weekday())).isoformat())
    return sorted(out)


def part_week(sheet_last: dt.date) -> str:
    """The week the *workbook* ends in - deviation 3.

    The sheet's, not the database's. Those were the same week while the
    spreadsheet was still being kept up; now that the dashboard has taken over
    they are not, and it is the sheet's last week that is averaged over a
    seven-cell window it has only partly filled.
    """
    return (sheet_last - dt.timedelta(days=sheet_last.weekday())).isoformat()


def last_reading(wb) -> dt.date:
    """The last day the Data sheet has a weight for."""
    latest = None
    for row in wb["Data"].iter_rows(min_row=3, max_col=2, values_only=True):
        when = as_date(row[0])
        if when is not None and row[1] is not None:
            latest = when
    if latest is None:
        raise SystemExit("The workbook's Data sheet holds no readings")
    return latest


def main() -> int:
    import openpyxl

    if not WORKBOOK.exists():
        print(f"Workbook not found: {WORKBOOK}")
        return 1
    if not queries.coverage()["days"]:
        print("No data - run: python -m core.excel_import --rebuild")
        return 1

    print(f"Reconciling against {WORKBOOK.name}")
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True, read_only=True)
    try:
        sheet_last = last_reading(wb)
        last_db = dt.date.fromisoformat(queries.coverage()["last_day"])
        print(f"  the workbook stops at {sheet_last:%d/%m/%Y}; "
              f"the database runs to {last_db:%d/%m/%Y}"
              + (f" — {(last_db - sheet_last).days} day(s) further on, so the "
                 f"period it stops in is compared knowing that"
                 if last_db > sheet_last else ""))

        # --- the daily figures, and the two derived masses -------------------
        compare(wb, "Daily averages", "Summary", queries.series("daily"),
                expect_absent=DROPPED)

        print("\nDerived masses (Data!T and Data!U)")
        daily = {r["period"]: r for r in queries.series("daily")}
        checked = wrong = 0
        for row in wb["Data"].iter_rows(min_row=3, max_col=21, values_only=True):
            when = as_date(row[0])
            if when is None or row[19] is None:
                continue
            mine = daily.get(when.isoformat())
            if mine is None:
                continue
            for column, key in ((19, "body_fat_kg"), (20, "muscle_kg")):
                checked += 1
                wrong += abs(float(mine[key]) - float(row[column])) > TOL
        check(f"{checked} masses compared", wrong, 0)

        # --- weekly averages read Data directly, so only the part-week differs
        compare(wb, "Weekly averages", "Weekly Results", queries.series("weekly"),
                expect_differs=[part_week(sheet_last)])

        # --- everything below is built on Summary, so it inherits deviation 1
        compare(wb, "Monthly averages", "Monthly average",
                queries.series("monthly"),
                expect_differs=affected_months(sheet_last))

        compare(wb, "Rolling 7-day change", "Weekly changes - Daily rolling",
                queries.rolling_change(7), expect_differs=affected_rolling())

        compare(wb, "Weekly average of the rolling change",
                "Weekly changes - Weekly average", queries.weekly_average_change(),
                expect_differs=sorted({*affected_weeks(),
                                       part_week(sheet_last)}))

        # The monthly change sheet labels its rows "August-2022", not a date.
        print("\nMonth-on-month change (Monthly average change)")
        mine = {}
        for row in queries.changes("monthly"):
            when = dt.date.fromisoformat(row["period"])
            mine[f"{when:%B-%Y}"] = (row["period"], row)
        differs, checked = [], 0
        for row in wb["Monthly average change"].iter_rows(
                min_row=3, max_col=7, values_only=True):
            entry = mine.get(str(row[0]))
            if entry is None:
                continue
            checked += 1
            if values_match(entry[1], row):
                differs.append(entry[0])
        check(f"{checked} months compared", checked > 0, True)
        check("months that differ", sorted(differs),
              affected_month_changes(sheet_last))
    finally:
        wb.close()

    print()
    if failures:
        print(f"FAILED: {len(failures)} check(s): {', '.join(failures)}")
        return 1
    print("Every figure reconciles with the workbook, "
          "deviations included and accounted for.")
    return 0


if __name__ == "__main__":
    sys.path.insert(0, str(Path(__file__).resolve().parent))
    sys.exit(main())
