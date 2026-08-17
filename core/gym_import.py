"""Import the gym programme workbook into the workout tables.

    python -m core.gym_import                 # add or update the plan
    python -m core.gym_import --rebuild       # delete it first and start again
    python -m core.gym_import --name "..."    # store it under another name

Reads `excel_versions/2026 Gym Programme.xlsx` strictly read-only.

Sheet layout
------------
**Programme Overview** holds the parts that are true of the whole plan:

    A5:B8     the 1RMs - four lifts, in kilograms
    A11:F14   the phases: weeks, focus, set scheme, working %, accessories
    A18:E24   the rotating pairings, which give each week its A/B type
    A28:F31   target weights: warm-up and working percentages per phase

**Tracker** is a tick per (week, workout): Workout A is session 1 and Workout B
is session 2.

**Week NN** is one sheet per week, and the shape it repeats is:

    WEEK n  ·  Phase 1 - Hypertrophy ...        row 1, the phase comes from here
      S1: Bench Press + Squats  |  S2: ...      row 2, a summary of row 4 onwards
    Exercise | Set Type | Set # | Reps | % 1RM | Weight (kg) | Rest | Notes
    SESSION 1  -  Bench Press + Squats          a session starts
      Bench Press                               an exercise starts
              | Warm-Up | W1 | 5 | 0.5 | 47.5 | 60s |          one set per row
    ...
      WEEK n COACHING NOTES                     the notes block, to week.note
      * ...

A row is a set if column B is filled in. Otherwise column A names something: a
session, an exercise, or a line of the notes. That test is what makes the parser
indifferent to the blank spacer rows and to the accessories, which repeat their
own name on their first set row as well as in the header above it.

What the workbook computes and this does not import
---------------------------------------------------
The Weight (kg) column, wherever the % 1RM column is filled in. It is
`% x 1RM` rounded to 2.5 kg and nothing else, and importing a derived value is
how a database ends up disagreeing with itself - the same rule the run tracker
follows for pace. The import checks every one of those cells against the
percentage instead, and says so in the summary it prints: either every weight
reconciles, or it lists the ones that do not, rather than silently preferring
one of the two.
"""
from __future__ import annotations

import argparse
import re
from pathlib import Path

import config
from core import db, workout_mutations as wm, workout_queries as wq, workouts

# openpyxl is imported inside run_import(). Importing this module has to stay
# cheap: the Flask front-end runs on a NAS with ~150 MB free and never imports
# a workbook.

# Column positions, 0-indexed into the tuple iter_rows hands back.
EXERCISE, SET_TYPE, SET_NUM, REPS, PERCENT, WEIGHT, REST, NOTE = range(8)

# What the sheet calls each kind of set.
SET_TYPES = {"warm-up": "warmup", "warmup": "warmup", "working": "working",
             "accessory": "accessory"}

# Weight cells that are words rather than numbers.
BODYWEIGHT = "bodyweight"
CHOOSE = ("choose weight", "light weight", "choose", "bodyweight or assisted")

# '+5 kg added', '+12.5 kg added'
ADDED = re.compile(r"^\+\s*([\d.]+)\s*kg", re.I)

# Lines in the notes block. The workbook uses a bullet, and an info glyph for the
# one-off explanation in the deload week.
NOTE_MARKERS = ("•", "‣", "-", "ℹ", "*")

# 'WEEK 7  ·  Phase 2 - Strength ...' / 'WEEK 19  ·  DELOAD  ·  ...'
PHASE_IN_TITLE = re.compile(r"(Phase\s*\d+|DELOAD)", re.I)


class GymImportError(RuntimeError):
    """The workbook is not shaped the way this importer expects."""


# --------------------------------------------------------------------------- #
# Reading the workbook
# --------------------------------------------------------------------------- #
def _text(value) -> str:
    return "" if value is None else " ".join(str(value).split())


def _number(value):
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = _text(value).replace(",", "")
    try:
        return float(text)
    except ValueError:
        return None


def _overview(ws) -> dict:
    """The 1RMs, the phases and the week -> A/B mapping."""
    maxes: dict = {}
    for row in range(5, 30):
        name = _text(ws.cell(row, 1).value)
        value = _number(ws.cell(row, 2).value)
        if name and value and name.lower() not in ("phase", "weeks"):
            if name.lower().startswith("phase"):
                break
            maxes[name] = value

    phases: list = []
    for row in range(11, 20):
        name = _text(ws.cell(row, 1).value)
        if not name.lower().startswith(("phase", "deload")):
            continue
        phases.append({
            "name": name,
            "weeks_text": _text(ws.cell(row, 2).value),
            "focus": _text(ws.cell(row, 3).value),
            "scheme": _text(ws.cell(row, 4).value),
            "working_text": _text(ws.cell(row, 5).value),
            "accessories": _text(ws.cell(row, 6).value),
        })

    # The target-weights table, which is where the percentages actually live.
    targets: dict = {}
    for row in range(28, 40):
        name = _text(ws.cell(row, 1).value)
        if not name:
            continue
        warmups = [v for v in (_number(ws.cell(row, col).value)
                               for col in (2, 3)) if v]
        working = [v for v in (_number(ws.cell(row, col).value)
                               for col in (4, 5, 6)) if v]
        if warmups or working:
            targets[name] = {"warmup": warmups, "working": working}

    # Weeks -> A or B, from the pairings table.
    #
    # Stopped at the first row that is not one. The target-weights table sits
    # four rows below with "Phase 1" in column A and 0.5 in column B, and a
    # looser scan reads that as "week 1 is of type 0.5" - which it did, until
    # the first import printed it.
    cycle: dict = {}
    for row in range(18, 40):
        weeks_text = _text(ws.cell(row, 1).value)
        kind = _text(ws.cell(row, 2).value)
        if not weeks_text:
            continue
        if not re.fullmatch(r"[\d\s,]+", weeks_text):
            break
        if not kind or len(kind) > 2 or kind in ("-", "–", "—"):
            continue
        for number in re.findall(r"\d+", weeks_text):
            cycle[int(number)] = kind

    return {"maxes": maxes, "phases": phases, "targets": targets,
            "cycle": cycle}


def _tracker(ws) -> dict:
    """week number -> [session numbers ticked]. Workout A is session 1."""
    ticked: dict = {}
    header_row = None
    for row in range(1, 6):
        if _text(ws.cell(row, 2).value).lower() == "week":
            header_row = row
            break
    if header_row is None:
        return ticked

    columns = {}
    for col in range(3, 12):
        label = _text(ws.cell(header_row, col).value)
        if not label:
            break
        # 'Workout A' is the first session of the week, 'Workout B' the second.
        columns[col] = len(columns) + 1

    for row in range(header_row + 1, ws.max_row + 1):
        number = _number(ws.cell(row, 2).value)
        if not number:
            continue
        done = [session for col, session in columns.items()
                if ws.cell(row, col).value is True]
        if done:
            ticked[int(number)] = done
    return ticked


def _week_sheet(ws) -> dict:
    """One week: its phase, its sessions, and the notes block at the bottom."""
    title = _text(ws.cell(1, 1).value)
    match = re.search(r"WEEK\s*(\d+)", title, re.I)
    if not match:
        raise GymImportError(f"{ws.title!r}: row 1 does not name a week "
                             f"({title[:60]!r})")
    number = int(match.group(1))
    phase_match = PHASE_IN_TITLE.search(title)
    phase = phase_match.group(1).title() if phase_match else None
    label = "Deload" if "deload" in title.lower() else None

    sessions: list = []
    notes: list = []
    current_session = None
    current_exercise = None
    in_notes = False

    for values in ws.iter_rows(min_row=4, max_col=8, values_only=True):
        first = _text(values[EXERCISE])
        set_type = _text(values[SET_TYPE]).lower()

        if set_type in SET_TYPES:
            if current_exercise is None:
                raise GymImportError(
                    f"{ws.title!r}: a {set_type} set with no exercise above it")
            current_exercise["sets"].append(_read_set(values, set_type))
            continue

        if not first:
            continue

        upper = first.upper()
        if upper.startswith("SESSION"):
            in_notes = False
            current_session = {"name": _session_name(first),
                               "number": len(sessions) + 1, "exercises": []}
            sessions.append(current_session)
            current_exercise = None
            continue

        if "COACHING NOTES" in upper or "GUIDANCE" in upper:
            in_notes = True
            continue

        if in_notes or first.startswith(NOTE_MARKERS):
            notes.append(first.lstrip("".join(NOTE_MARKERS) + " ").strip())
            continue

        if first.lower() == "exercise":      # a repeated heading row
            continue

        # Anything left naming something is an exercise header. The accessories
        # repeat their name on their first set row too, which the set-row branch
        # above has already consumed - so seeing the same name twice in a row is
        # normal and the second one never reaches here.
        if current_session is None:
            continue
        if current_exercise is not None and current_exercise["name"] == first:
            continue
        current_exercise = {"name": first, "sets": []}
        current_session["exercises"].append(current_exercise)

    return {"number": number, "phase": phase, "label": label,
            "sessions": sessions,
            "note": "\n".join(line for line in notes if line) or None}


def _session_name(text: str) -> str:
    """'SESSION 1  -  Bench Press + Squats' -> 'Bench Press + Squats'."""
    parts = re.split(r"[-–—]", text, maxsplit=1)
    return _text(parts[1]) if len(parts) > 1 else _text(text)


def _read_set(values, set_type: str) -> dict:
    """One row of the sheet as the fields save_session() wants."""
    reps_raw = values[REPS]
    reps_text = _text(reps_raw) or ""
    percent = _number(values[PERCENT])
    weight_raw = values[WEIGHT]
    weight_number = _number(weight_raw)
    weight_text = _text(weight_raw).lower()

    load: dict = {"load_mode": "choose"}
    if percent:
        # The percentage is the instruction; the kilograms beside it are its
        # consequence, and are checked rather than imported.
        load = {"load_mode": "percent", "percent_1rm": percent}
    elif weight_text.startswith(BODYWEIGHT) or ADDED.match(weight_text):
        added = ADDED.match(weight_text)
        load = {"load_mode": "bodyweight",
                "added_kg": float(added.group(1)) if added else None}
    elif weight_number is not None:
        load = {"load_mode": "explicit", "weight_kg": weight_number}
    elif weight_text and weight_text not in CHOOSE:
        # An unexpected word. Kept as an unprescribed set with the word as the
        # cue, so the run does not stop and the oddity is visible.
        load = {"load_mode": "choose"}

    return {
        "set_type": SET_TYPES[set_type],
        "reps": reps_text,
        "rest": _text(values[REST]),
        "cue": _text(values[NOTE]),
        "sheet_weight": weight_number,
        "sheet_weight_text": _text(weight_raw),
        "per_side": "each" in reps_text.lower(),
        **load,
    }


# --------------------------------------------------------------------------- #
# Writing it
# --------------------------------------------------------------------------- #
def run_import(source=None, name: str | None = None, rebuild: bool = False,
               db_path=None) -> dict:
    """Fold the workbook into one plan. Returns a summary of what it did."""
    import openpyxl

    source = Path(source or config.GYM_XLSX)
    if not source.exists():
        raise FileNotFoundError(f"Gym workbook not found: {source}")
    plan_name = name or source.stem

    db.init_db(db_path)

    wb = openpyxl.load_workbook(source, read_only=False, data_only=True)
    try:
        if "Programme Overview" not in wb.sheetnames:
            raise GymImportError(
                f"{source.name} has no 'Programme Overview' sheet - "
                f"found {wb.sheetnames[:4]}")
        overview = _overview(wb["Programme Overview"])
        ticked = _tracker(wb["Tracker"]) if "Tracker" in wb.sheetnames else {}
        weeks = [_week_sheet(wb[title]) for title in wb.sheetnames
                 if title.lower().startswith("week")]
    finally:
        wb.close()

    weeks.sort(key=lambda item: item["number"])
    if not weeks:
        raise GymImportError(f"{source.name} has no week sheets")

    existing = wq.plan_by_name(plan_name)
    if existing and not rebuild:
        return {"plan": plan_name, "plan_id": existing["id"], "skipped": True,
                "reason": "already imported - pass --rebuild to replace it",
                "weeks": existing["weeks"], "sessions": existing["sessions"]}
    if existing:
        wm.delete_plan(existing["id"])

    # Anything the sheet names that the catalogue does not know yet. Added
    # rather than refused: the workbook is where this vocabulary comes from.
    added_exercises = _register_exercises(weeks)

    plan = wm.save_plan({"name": plan_name, "rounding_kg": 2.5,
                         "note": f"Imported from {source.name}"},
                        source="xlsx")
    plan_id = plan["id"]

    for lift, one_rm in overview["maxes"].items():
        row = wq.exercise_by_name(lift)
        if row is not None:
            wm.set_max(plan_id, row["id"], one_rm)

    phase_ids = _write_phases(plan_id, overview)

    counts = {"weeks": 0, "sessions": 0, "exercises": 0, "sets": 0}
    mismatches: list = []
    for week in weeks:
        created = wm.save_week(plan_id, {
            "number": week["number"],
            "label": week["label"],
            "phase_id": phase_ids.get((week["phase"] or "").title()),
            "cycle_type": overview["cycle"].get(week["number"]),
            "note": week["note"],
        })
        counts["weeks"] += 1
        for session in week["sessions"]:
            exercises = [{
                "exercise_id": wq.exercise_by_name(item["name"])["id"],
                "sets": item["sets"],
            } for item in session["exercises"] if wq.exercise_by_name(item["name"])]
            if not exercises:
                continue
            saved = wm.save_session(
                created["id"],
                {"number": session["number"], "name": session["name"]},
                exercises)
            counts["sessions"] += 1
            counts["exercises"] += saved["exercises"]
            counts["sets"] += saved["sets"]
        mismatches += _check_week(plan_id, created["id"], week)

    ticks = 0
    for number, session_numbers in ticked.items():
        week_row = wq.week_number(plan_id, number)
        if week_row is None:
            continue
        by_number = {item["number"]: item["id"]
                     for item in wq.sessions(week_id=week_row["id"])}
        for session_number in session_numbers:
            if session_number in by_number:
                wm.tick_session(by_number[session_number], True)
                ticks += 1

    with db.transaction(db_path) as conn:
        db.log(conn, "import", "plans", str(plan_id),
               f"{source.name}: {counts['weeks']} weeks, "
               f"{counts['sessions']} sessions, {counts['sets']} sets")

    return {
        "plan": plan_name,
        "plan_id": plan_id,
        "skipped": False,
        **counts,
        "one_rms": overview["maxes"],
        "phases": len(phase_ids),
        "sessions_ticked": ticks,
        "exercises_added": added_exercises,
        "weight_mismatches": mismatches,
    }


def _register_exercises(weeks) -> list:
    """Add any movement the sheet names that the catalogue has not got."""
    added = []
    for week in weeks:
        for session in week["sessions"]:
            for item in session["exercises"]:
                if wq.exercise_by_name(item["name"]) is None:
                    per_side = any(s.get("per_side") for s in item["sets"])
                    bodyweight = any(s.get("load_mode") == "bodyweight"
                                     for s in item["sets"])
                    wm.save_exercise({
                        "name": item["name"],
                        "reps_mode": "per_side" if per_side else "total",
                        "is_bodyweight": bodyweight,
                        "note": "Added by the gym workbook import",
                    })
                    added.append(item["name"])
    return added


def _write_phases(plan_id: int, overview: dict) -> dict:
    """The phase table plus the target-weights table, which pair up by name."""
    ids: dict = {}
    for phase in overview["phases"]:
        targets = overview["targets"].get(phase["name"], {})
        scheme = phase["scheme"]
        working_sets, working_reps = _scheme(scheme)
        accessory_sets = _first_int(re.search(r"(\d+)\s*sets",
                                              phase["accessories"], re.I))
        made = wm.save_phase(plan_id, {
            "name": phase["name"],
            "focus": phase["focus"] or phase["weeks_text"] or None,
            "warmup_pcts": targets.get("warmup") or [],
            "working_pcts": targets.get("working") or [],
            "working_sets": working_sets,
            "working_reps": working_reps,
            "accessory_sets": accessory_sets,
            "accessory_reps": _reps_from(phase["accessories"]),
            "rest_warmup": config.DEFAULT_REST["warmup"],
            "rest_working": config.DEFAULT_REST["working"],
            "rest_accessory": config.DEFAULT_REST["accessory"],
        })
        ids[phase["name"].title()] = made["id"]
    return ids


def _first_int(match):
    return int(match.group(1)) if match else None


def _scheme(text: str) -> tuple:
    """(working sets, reps) from a phase's set scheme, however it is written.

    Two spellings in the workbook, because the peaking phase changes shape every
    fortnight and the others do not:

        "2 warm-up sets / 5 working sets x 10 reps"     -> (5, '10')
        "Wks 13-14: 5x3 @ ~87% / Wks 15-16: 5x2 ..."    -> (5, '3')

    The second takes the first pairing it finds. A phase is only ever a set of
    defaults the session builder pre-fills from, and the fortnight-by-fortnight
    detail is in `focus` where it can be read.
    """
    sets = _first_int(re.search(r"(\d+)\s*working", text, re.I))
    reps = _reps_from(text)
    if sets is None or reps is None:
        pair = re.search(r"(\d+)\s*[x×]\s*(\d+)(?!\s*reps)", text, re.I)
        if pair:
            sets = sets if sets is not None else int(pair.group(1))
            reps = reps if reps is not None else pair.group(2)
    return sets, reps


def _reps_from(text: str):
    """'5 working sets x 10 reps' -> '10'; '3 sets x 10-12 reps' -> '10-12'."""
    match = re.search(r"[x×]\s*(\d+\s*[-–]\s*\d+|\d+)\s*reps", text,
                      re.I)
    if not match:
        return None
    try:
        low, high = workouts.parse_reps(match.group(1))
    except workouts.InvalidWorkout:
        return None
    return f"{low}-{high}" if high else f"{low}"


def _check_week(plan_id: int, week_id: int, week: dict) -> list:
    """Compare the sheet's Weight column against the percentage it came from.

    The importer does not load that column - it is derived - so this is the only
    thing keeping the two honest. A mismatch means either the sheet was edited
    without recalculating or the rounding rule is not what it looks like, and
    both are worth being told about rather than reconciled silently.
    """
    wanted: dict = {}
    for session in week["sessions"]:
        for item in session["exercises"]:
            for order, entry in enumerate(item["sets"]):
                if entry.get("load_mode") == "percent" and entry.get("sheet_weight"):
                    key = (item["name"], entry["set_type"])
                    wanted.setdefault(key, []).append(entry["sheet_weight"])

    seen: dict = {}
    problems = []
    for row in db.rows(
            "SELECT exercise_name, set_type, position, percent_1rm, "
            "prescribed_kg FROM v_exercise_sets WHERE week_id = ? "
            "AND load_mode = 'percent' ORDER BY exercise_position, set_type, "
            "position", (week_id,)):
        key = (row["exercise_name"], row["set_type"])
        index = seen.get(key, 0)
        seen[key] = index + 1
        expected = wanted.get(key, [])
        if index >= len(expected) or row["prescribed_kg"] is None:
            continue
        if abs(row["prescribed_kg"] - expected[index]) > 0.001:
            problems.append({
                "week": week["number"], "exercise": row["exercise_name"],
                "set": f"{row['set_type']} {row['position']}",
                "percent": row["percent_1rm"],
                "sheet_kg": expected[index],
                "worked_out_kg": row["prescribed_kg"],
            })
    return problems


# --------------------------------------------------------------------------- #
# Command line
# --------------------------------------------------------------------------- #
def main() -> None:
    parser = argparse.ArgumentParser(
        description="Import the gym programme workbook into the workout tables")
    parser.add_argument("--source", type=Path, default=config.GYM_XLSX)
    parser.add_argument("--name", default=None,
                        help="plan name (defaults to the workbook's filename)")
    parser.add_argument("--rebuild", action="store_true",
                        help="delete the plan of that name first")
    args = parser.parse_args()

    result = run_import(args.source, name=args.name, rebuild=args.rebuild)
    if result.get("skipped"):
        print(f"'{result['plan']}' {result['reason']}")
        return

    print(f"Imported '{result['plan']}' (plan {result['plan_id']})")
    for key in ("weeks", "sessions", "exercises", "sets", "phases",
                "sessions_ticked"):
        print(f"  {key.replace('_', ' '):<18} {result[key]}")
    print(f"  {'1RMs':<18} " + ", ".join(
        f"{lift} {workouts.fmt_kg(kg)}kg"
        for lift, kg in result["one_rms"].items()))
    if result["exercises_added"]:
        print(f"  exercises added    {', '.join(result['exercises_added'])}")
    problems = result["weight_mismatches"]
    if problems:
        print(f"\n  {len(problems)} weight(s) in the sheet do not match the "
              f"percentage beside them:")
        for row in problems[:20]:
            print(f"    week {row['week']:>2}  {row['exercise']:<28} "
                  f"{row['set']:<12} {row['percent'] * 100:g}%  "
                  f"sheet {row['sheet_kg']}  worked out "
                  f"{row['worked_out_kg']}")
    else:
        print("\n  every weight in the sheet matches the percentage beside it")


if __name__ == "__main__":
    main()
