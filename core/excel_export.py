"""Write an .xlsx snapshot of the database.

The dashboard is the system of record, but a spreadsheet snapshot means the data
is never locked inside an app - useful as a backup, and it reproduces the tabs
the original workbook had, so anything you still want to do in Excel you can.

Uses openpyxl directly rather than pandas. pandas costs 85 MB of RAM to import,
which the NAS cannot spare; openpyxl is pure Python and is imported lazily so
it costs nothing until an export is actually requested.
"""
from __future__ import annotations

import datetime as dt
from pathlib import Path

import config
from core import db, queries

# The workbook's own tab names, so the export reads like the original.
SHEETS = [
    ("Data",                    lambda: _raw_readings()),
    ("Summary",                 lambda: queries.series("daily")),
    ("Weekly Results",          lambda: queries.series("weekly")),
    ("Monthly average",         lambda: queries.series("monthly")),
    ("Daily changes",           lambda: queries.changes("daily")),
    ("Weekly rolling changes",  lambda: queries.rolling_change(7)),
    ("Weekly average changes",  lambda: queries.weekly_average_change()),
    ("Monthly average change",  lambda: queries.changes("monthly")),
]

HEADINGS = {
    "period": "Date", "previous": "Compared with", "days": "Days",
    "estimated_days": "Estimated", "estimated": "Estimated",
    "readings": "Weigh-ins", "slot": "Weigh-in",
}


def _heading(column: str) -> str:
    if column in HEADINGS:
        return HEADINGS[column]
    return config.LABELS.get(column, column.replace("_", " ").capitalize())


def _raw_readings() -> list:
    """The Data sheet: both weigh-ins, as entered."""
    return db.rows(
        f"SELECT day AS period, slot, {', '.join(config.METRIC_KEYS)}, "
        "estimated, note FROM readings ORDER BY day, slot")


def export(path: Path | None = None) -> Path:
    from openpyxl import Workbook  # lazy: keeps the Flask baseline small
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    if path is None:
        config.EXPORT_DIR.mkdir(parents=True, exist_ok=True)
        stamp = dt.datetime.now().strftime("%Y%m%d_%H%M")
        path = config.EXPORT_DIR / f"Weigh_in_Tracker_{stamp}.xlsx"

    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet

    bold = Font(bold=True)
    for name, fetch in SHEETS:
        rows = fetch()
        ws = wb.create_sheet(name[:31])
        if not rows:
            ws.append(["No data"])
            continue

        columns = list(rows[0])
        ws.append([_heading(c) for c in columns])
        for cell in ws[1]:
            cell.font = bold

        for row in rows:
            ws.append([row.get(c) for c in columns])

        ws.freeze_panes = "A2"
        for index, column in enumerate(columns, start=1):
            width = max(len(_heading(column)) + 2, 11)
            ws.column_dimensions[get_column_letter(index)].width = min(width, 24)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)

    with db.transaction() as conn:
        db.log(conn, "export", "workbook", None, str(path))
    return path


if __name__ == "__main__":
    print("Written:", export())
