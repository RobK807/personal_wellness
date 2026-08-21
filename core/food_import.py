"""Import the food planner workbook, and round-trip the diary through a CSV.

    python -m core.food_import --catalogue        # the Food sheet -> foods
    python -m core.food_import --export           # the diary -> a CSV to edit
    python -m core.food_import --load FILE        # that CSV -> the database
    python -m core.food_import --catalogue --export

Reads `Food Planner v0.1.xlsx` strictly read-only.

Two things come out of that workbook and they arrive by different routes.

**The catalogue** goes straight in. The Food sheet holds three lists side by
side - Items at B, Meals at N, Recipes at Z - each eight columns of Name,
Grouping, Portion, Units and the four macros. 183 rows, all of them numeric,
nothing to interpret.

**The diary does not.** Food_Diary is 49 week blocks of seven columns, each
holding seven days of twenty rows, and every line of it is a *rendered string* -
"All Real bar - 1 Bar" - rather than a food and a quantity. 287 distinct strings
against 182 catalogue entries, and the difference is spelling, portions written
into names, and things eaten once. Guessing at that would quietly rewrite two
years of history, so it is exported to a CSV to be corrected by hand and read
back with `--load`.

The export does the guessing it can and shows its working: it splits each string
on the last " - number units" it finds, and matches the remainder against the
catalogue. Both are offered as separate columns beside the untouched original,
so a wrong guess is visible rather than baked in.

Sodium and sugar are not imported. The workbook has columns for them and no data
in them.
"""
from __future__ import annotations

import argparse
import csv
import datetime as dt
import re
from pathlib import Path

import config
from core import db, food, food_mutations as fm, food_queries as fq

# openpyxl is imported inside the functions that need it. Importing this module
# has to stay cheap: the Flask front-end runs on a NAS with ~150 MB free.

# The Food sheet: three lists, each eight columns wide, starting at these
# 1-based columns. Read from the sheet's own header row rather than assumed.
CATALOGUE_BLOCKS = {"Items": 2, "Meals": 14, "Recipes": 26}
CATALOGUE_FIRST_ROW = 4

# Food_Diary's shape. A week block is seven columns; a day is twenty rows,
# seven of them per block starting at DAY_FIRST_ROW.
DIARY_FIRST_COL = 10          # J
WEEK_STRIDE = 7
DAY_FIRST_ROW = 6
DAY_STRIDE = 20
ENTRIES_PER_DAY = 16          # the rows between the day's header and its totals

# 'W/C 03/08/2026'
WEEK_HEADER = re.compile(r"W/C\s*(\d{2})/(\d{2})/(\d{4})", re.I)

CSV_COLUMNS = ["date", "weekday", "meal", "position", "recorded", "name",
               "quantity", "units", "calories", "carbs", "fat", "protein",
               "matched_food", "matched_list", "note"]


class FoodImportError(RuntimeError):
    """The workbook is not shaped the way this importer expects."""


def _text(value) -> str:
    return "" if value is None else " ".join(str(value).split())


def _number(value):
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(_text(value).replace(",", ""))
    except ValueError:
        return None


# --------------------------------------------------------------------------- #
# The catalogue
# --------------------------------------------------------------------------- #
def read_catalogue(source: Path) -> list:
    """The three food lists, as rows ready for the foods table."""
    import openpyxl

    wb = openpyxl.load_workbook(source, data_only=True, read_only=False)
    try:
        if "Food" not in wb.sheetnames:
            raise FoodImportError(f"{source.name} has no 'Food' sheet")
        ws = wb["Food"]
        rows = []
        for list_name, col in CATALOGUE_BLOCKS.items():
            for row in range(CATALOGUE_FIRST_ROW, ws.max_row + 1):
                name = _text(ws.cell(row, col).value)
                if not name:
                    continue
                macros = [_number(ws.cell(row, col + offset).value)
                          for offset in (4, 5, 6, 7)]
                if any(value is None for value in macros):
                    continue      # a half-filled row is not a food
                rows.append({
                    "list": list_name,
                    "name": name,
                    "grouping": _text(ws.cell(row, col + 1).value) or None,
                    "portion": _number(ws.cell(row, col + 2).value) or 1.0,
                    "units": _text(ws.cell(row, col + 3).value) or "Portion",
                    "calories": macros[0], "carbs": macros[1],
                    "fat": macros[2], "protein": macros[3],
                })
    finally:
        wb.close()
    return rows


def import_catalogue(source=None, db_path=None) -> dict:
    """Load the catalogue. Existing rows are updated, nothing is deleted."""
    source = Path(source or config.FOOD_XLSX)
    if not source.exists():
        raise FileNotFoundError(f"Food workbook not found: {source}")
    db.init_db(db_path)
    rows = read_catalogue(source) + extra_foods()
    result = fm.load_foods(rows, source="xlsx")
    return {"source": source.name, "read": len(rows), **result}


def extra_foods() -> list:
    """The catalogue rows config knows about and the workbook does not.

    Loaded with the sheet rather than after it so that a database rebuilt from
    scratch resolves every line of the corrected diary - see the note on
    config.FOOD_EXTRA_FOODS. They go through the same fm.load_foods() as the
    sheet, so one edited in the app is updated rather than duplicated.
    """
    keys = ("list", "name", "grouping", "portion", "units",
            "calories", "carbs", "fat", "protein", "note")
    return [dict(zip(keys, values)) for values in config.FOOD_EXTRA_FOODS]


# --------------------------------------------------------------------------- #
# The diary
# --------------------------------------------------------------------------- #
def read_diary(source: Path) -> dict:
    """Every diary line in the workbook, with the week it came from.

    Returns the entries plus what was skipped, because the two duplicated week
    blocks and the empty ones are worth being told about rather than silently
    folded together.
    """
    import openpyxl

    wb = openpyxl.load_workbook(source, data_only=True, read_only=False)
    try:
        if "Food_Diary" not in wb.sheetnames:
            raise FoodImportError(f"{source.name} has no 'Food_Diary' sheet")
        ws = wb["Food_Diary"]

        blocks = []
        for col in range(1, ws.max_column + 1):
            match = WEEK_HEADER.search(_text(ws.cell(1, col).value))
            if match:
                monday = dt.date(int(match.group(3)), int(match.group(2)),
                                 int(match.group(1)))
                blocks.append((col, monday))

        entries, seen, duplicates = [], {}, []
        for col, monday in blocks:
            week = _read_week(ws, col, monday)
            if not week:
                continue
            if monday in seen:
                # Two blocks claim the same week. Keep the fuller one and say
                # so - picking silently would lose whichever half was newer.
                if len(week) > len(seen[monday]["entries"]):
                    duplicates.append({"week": monday.isoformat(),
                                       "kept": len(week),
                                       "dropped": len(seen[monday]["entries"])})
                    entries = [e for e in entries
                               if food.week_start(e["date"], 0) != monday]
                    seen[monday] = {"entries": week}
                    entries += week
                else:
                    duplicates.append({"week": monday.isoformat(),
                                       "kept": len(seen[monday]["entries"]),
                                       "dropped": len(week)})
                continue
            seen[monday] = {"entries": week}
            entries += week
    finally:
        wb.close()

    entries.sort(key=lambda row: (row["date"], config.MEALS.index(row["meal"])
                                  if row["meal"] in config.MEALS else 9,
                                  row["position"]))
    return {"entries": entries, "weeks": len(seen), "blocks": len(blocks),
            "duplicates": duplicates}


def _read_week(ws, col: int, monday: dt.date) -> list:
    """One week block: seven days of up to sixteen lines each."""
    out = []
    for index in range(7):
        start = DAY_FIRST_ROW + DAY_STRIDE * index
        day = monday + dt.timedelta(days=index)
        meal, position = None, 0
        for row in range(start + 2, start + 2 + ENTRIES_PER_DAY):
            label = _text(ws.cell(row, col).value)
            if label:
                meal, position = label, 0
            name = _text(ws.cell(row, col + 1).value)
            calories = _number(ws.cell(row, col + 2).value)
            if not name or calories is None:
                continue
            position += 1
            out.append({
                "date": day,
                "meal": meal or "Snacks",
                "position": position,
                "recorded": name,
                "calories": calories,
                "carbs": _number(ws.cell(row, col + 3).value) or 0.0,
                "fat": _number(ws.cell(row, col + 4).value) or 0.0,
                "protein": _number(ws.cell(row, col + 5).value) or 0.0,
            })
    return out


def export_diary(source=None, path=None, db_path=None) -> dict:
    """Write the diary to a CSV for correcting by hand.

    The catalogue is consulted for the match column, so `--catalogue` first
    gives a more useful file - but it is not required, and an empty match column
    simply means nothing was suggested.
    """
    source = Path(source or config.FOOD_XLSX)
    if not source.exists():
        raise FileNotFoundError(f"Food workbook not found: {source}")
    path = Path(path or config.DIARY_CSV)
    path.parent.mkdir(parents=True, exist_ok=True)

    db.init_db(db_path)
    read = read_diary(source)
    lookup = _catalogue_lookup()

    matched = flagged = 0
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=CSV_COLUMNS)
        writer.writeheader()
        for entry in read["entries"]:
            split = food.split_description(entry["recorded"])
            hit = lookup.get(split["name"].casefold()) \
                or lookup.get(entry["recorded"].casefold())
            matched += hit is not None
            note = _suspect(entry, hit)
            flagged += bool(note)
            writer.writerow({
                "date": entry["date"].isoformat(),
                "weekday": food.day_name(entry["date"]),
                "meal": entry["meal"],
                "position": entry["position"],
                "recorded": entry["recorded"],
                "name": split["name"],
                "quantity": split["quantity"] if split["quantity"] is not None
                            else "",
                "units": split["units"] or "",
                "calories": entry["calories"], "carbs": entry["carbs"],
                "fat": entry["fat"], "protein": entry["protein"],
                "matched_food": hit["name"] if hit else "",
                "matched_list": hit["list"] if hit else "",
                "note": note,
            })

    days = {entry["date"] for entry in read["entries"]}
    return {
        "path": path,
        "entries": len(read["entries"]),
        "days": len(days),
        "weeks": read["weeks"],
        "blocks": read["blocks"],
        "duplicates": read["duplicates"],
        "matched": matched,
        "unmatched": len(read["entries"]) - matched,
        "flagged": flagged,
        "first_day": min(days).isoformat() if days else None,
        "last_day": max(days).isoformat() if days else None,
    }


def _suspect(entry: dict, match: dict | None) -> str:
    """A note on the rows worth a second look, written into the CSV.

    Two tells, both from the workbook rather than imagined. A macro outside the
    bounds the input form enforces is either a typo or a total that leaked into
    a line. And four identical macros is a formula filled across a row - which
    happened once, to a pizza recorded as 1584 of everything.
    """
    macros = [float(entry.get(key) or 0) for key in config.MACRO_KEYS]
    if len(set(macros)) == 1 and macros[0] > 0:
        hint = ""
        if match:
            hint = (" - the catalogue says "
                    + ", ".join(f"{config.MACRO_LABELS[key]} {match[key]:g}"
                                for key in config.MACRO_KEYS))
        return ("CHECK: all four macros are identical, which is a formula "
                "filled across" + hint)
    for key, value in zip(config.MACRO_KEYS, macros):
        low, high = config.FOOD_BOUNDS[key]
        if not low <= value <= high:
            return (f"CHECK: {config.MACRO_LABELS[key].lower()} of {value:g} is "
                    f"outside {low:g}-{high:g} and will be refused on load")
    return ""


def _catalogue_lookup() -> dict:
    """Lower-cased food name -> the catalogue row, for the match column."""
    out = {}
    for row in fq.foods(include_retired=True):
        out.setdefault(row["name"].casefold(), row)
    return out


def load_diary(path=None, db_path=None, replace: bool = False) -> dict:
    """Read a corrected CSV back in.

    Every row is parsed before anything is written, so a bad line stops the
    whole file rather than leaving half a diary loaded. `--replace` clears the
    days the file covers first, which is what makes a corrected export safe to
    load twice.
    """
    path = Path(path or config.DIARY_CSV)
    if not path.exists():
        raise FileNotFoundError(f"No diary CSV at {path}")
    db.init_db(db_path)

    lookup = {}
    for row in fq.foods(include_retired=True):
        lookup[(row["list"], row["name"].casefold())] = row["id"]
        lookup.setdefault((None, row["name"].casefold()), row["id"])

    rows, problems = [], []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        for number, raw in enumerate(csv.DictReader(handle), start=2):
            try:
                rows.append(_csv_row(raw, lookup))
            except (food.InvalidFood, ValueError, KeyError) as exc:
                problems.append({"line": number, "why": str(exc),
                                 "row": raw.get("recorded") or raw.get("name")})
    if problems:
        # Every bad line, not just the first: correcting a two-year diary is
        # one pass through a spreadsheet, and being told about them one at a
        # time turns that into twenty.
        shown = "\n".join(
            f"  line {row['line']}: {row['why']}"
            + (f"  ({row['row']})" if row["row"] else "")
            for row in problems[:20])
        more = (f"\n  ...and {len(problems) - 20} more"
                if len(problems) > 20 else "")
        raise FoodImportError(
            f"{len(problems)} row(s) in {path.name} could not be read, so "
            f"nothing was loaded:\n{shown}{more}")

    return {"path": path, **fm.load_entries(rows, replace=replace,
                                            source="csv")}


def _csv_date(raw: dict) -> dt.date:
    """The date of a CSV line, whichever way the spreadsheet wrote it.

    The export writes ISO. Excel does not keep it that way: opening the file in a
    UK locale and saving it turns every date into dd/mm/yyyy, which `as_date`
    rightly refuses - 03/08/2026 is either the 3rd of August or the 8th of March
    and nothing in the string says which.

    The `weekday` column beside it does. It is written by the export and Excel
    has no reason to touch it, so a candidate reading that disagrees with the
    named weekday is wrong and is discarded. dd/mm is tried before mm/dd because
    the workbook is British, but a file where only one reading fits its own
    weekday is decided by the file rather than by that assumption.
    """
    text = _text(raw.get("date"))
    named = _text(raw.get("weekday")).casefold()
    fallback = None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d/%m/%y", "%d-%m-%Y"):
        try:
            day = dt.datetime.strptime(text, fmt).date()
        except ValueError:
            continue
        if not named or config.WEEKDAY_NAMES[day.weekday()].casefold() == named:
            return day
        fallback = fallback or day
    if fallback is not None:
        # Parses, but no reading agrees with the weekday written next to it.
        # That is a row somebody has retyped, so take the date and say nothing:
        # the weekday column is a check, not a field anybody maintains.
        return fallback
    raise food.InvalidFood(f"'{raw.get('date')}' is not a date")


def _csv_row(raw: dict, lookup: dict) -> dict:
    """One CSV line, validated, as the mutation layer wants it."""
    day = _csv_date(raw)
    entry = food.parse_entry({
        "meal": raw.get("meal"),
        "name": raw.get("name") or raw.get("recorded"),
        "quantity": raw.get("quantity"),
        "units": raw.get("units"),
        **{key: raw.get(key) for key in config.MACRO_KEYS},
    })
    name = _text(raw.get("matched_food"))
    if name:
        list_name = _text(raw.get("matched_list")).title() or None
        food_id = lookup.get((list_name, name.casefold())) \
            or lookup.get((None, name.casefold()))
        if food_id is None:
            raise food.InvalidFood(
                f"'{name}' is not in the catalogue - check matched_food, or "
                f"clear it to keep the line as free text")
        entry["food_id"] = food_id
    position = raw.get("position")
    return {"day": day, "position": int(position) if position else None,
            "note": _text(raw.get("note")) or None, **entry}


# --------------------------------------------------------------------------- #
# Command line
# --------------------------------------------------------------------------- #
def main() -> None:
    parser = argparse.ArgumentParser(
        description="Import the food planner workbook")
    parser.add_argument("--source", type=Path, default=config.FOOD_XLSX)
    parser.add_argument("--catalogue", action="store_true",
                        help="load the Food sheet into the catalogue")
    parser.add_argument("--export", action="store_true",
                        help="write the diary to a CSV for correcting by hand")
    parser.add_argument("--load", type=Path, nargs="?", const=config.DIARY_CSV,
                        help="read a corrected CSV into the database")
    parser.add_argument("--csv", type=Path, default=config.DIARY_CSV,
                        help="where --export writes to")
    parser.add_argument("--replace", action="store_true",
                        help="with --load, clear the days the file covers first")
    args = parser.parse_args()

    if not (args.catalogue or args.export or args.load):
        parser.print_help()
        return

    if args.catalogue:
        result = import_catalogue(args.source)
        print(f"Catalogue from {result['source']}")
        for key in ("read", "added", "updated", "unchanged"):
            print(f"  {key:<12} {result[key]}")
        if result.get("by_list"):
            print("  by list      " + ", ".join(
                f"{name} {count}" for name, count in result["by_list"].items()))

    if args.export:
        result = export_diary(args.source, args.csv)
        print(f"\nDiary exported to {result['path']}")
        print(f"  {result['entries']:,} entries over {result['days']} days, "
              f"{result['first_day']} to {result['last_day']}")
        print(f"  {result['weeks']} weeks from {result['blocks']} blocks")
        print(f"  matched to the catalogue: {result['matched']:,}")
        print(f"  left for you to check   : {result['unmatched']:,}")
        if result["flagged"]:
            print(f"  flagged in the note col : {result['flagged']} "
                  f"(search the CSV for CHECK)")
        for row in result["duplicates"]:
            print(f"  duplicate week {row['week']}: kept the block with "
                  f"{row['kept']} entries, dropped the one with {row['dropped']}")
        print("\n  Correct `name`, `quantity`, `units` and `matched_food`, then:")
        print(f"    python -m core.food_import --load \"{result['path']}\" "
              f"--replace")

    if args.load:
        result = load_diary(args.load, replace=args.replace)
        print(f"\nLoaded {result['path']}")
        for key in ("days", "entries", "linked", "free_text"):
            print(f"  {key:<12} {result[key]}")


if __name__ == "__main__":
    main()
