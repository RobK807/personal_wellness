"""One-time (and re-runnable) import of the workbook into SQLite.

Reads `Weigh-in Tracker.xlsx` strictly read-only and rebuilds the database from
its Data sheet. Run it again whenever you want to re-baseline:

    python -m core.excel_import --rebuild

Sheet layout, all 1-indexed, established by inspecting the workbook:

  Data   row 1 groups the columns, row 2 holds the headings, data from row 3
      A  Date
      B-G  weigh-in 1: Weight, BMI, Body fat %, Skeletal muscle %,
                       RM Kcal, Visceral fat
      H-M  weigh-in 2: the same six
      N-S  the average of the two - formulas, so not imported
      T-U  body fat / muscle mass - formulas, so not imported

Only B-M are read. Everything from N rightwards is derived, and deriving it
here rather than importing it is the whole point of the exercise.

Which rows were interpolated
----------------------------
The workbook did not record that, so it is inferred: the scale reports weight,
BMI and the two percentages to one decimal place and RM Kcal and visceral fat
as whole numbers, so a stored 72.45 or 1667.5 cannot have been read off it and
must have come from the Adjustments sheet. That catches most of them - about
14% of the history - but not an interpolated day that happened to land on a
value the scale could have produced. It is a best-effort flag on imported
history, not a guarantee; everything entered through the dashboard from here on
is marked properly at the point of entry.
"""
from __future__ import annotations

import argparse
from pathlib import Path

import config
from core import db, metrics, mutations

# openpyxl is imported lazily inside run_import(). Importing this module must
# stay cheap: the Flask front-end runs on a NAS with ~150 MB of RAM free, and
# the workbook is normally imported once on a desktop rather than on the NAS.

SHEET = "Data"
FIRST_ROW = 3
DATE_COLUMN = 1
# Weigh-in 1 starts at column B, weigh-in 2 at column H, six metrics each.
SLOT_FIRST_COLUMN = {1: 2, 2: 8}

MAX_BLANK_RUN = 5  # stop scanning after this many consecutive dateless rows


def _looks_interpolated(values: dict) -> bool:
    """True if any value is finer than the scale can report. See the note above."""
    for key, value in values.items():
        if value is None:
            continue
        if abs(value - metrics.round_metric(key, value)) > 1e-9:
            return True
    return False


def _read_rows(ws) -> list:
    """Yield (date, {slot: values}) for every populated row of the Data sheet.

    Iterated rather than addressed cell by cell: openpyxl's read-only mode
    streams the sheet, so `ws.cell(r, c)` rescans it from the top every call.
    On 2,300 rows x 13 columns that is the difference between a second and
    several minutes.
    """
    out = []
    blanks = 0
    for row in ws.iter_rows(min_row=FIRST_ROW, max_col=13, values_only=True):
        raw_date = row[DATE_COLUMN - 1]
        if raw_date is None:
            blanks += 1
            if blanks > MAX_BLANK_RUN:
                break
            continue
        blanks = 0

        try:
            when = metrics.as_date(raw_date)
        except metrics.InvalidReading:
            continue

        slots = {}
        for slot, first in SLOT_FIRST_COLUMN.items():
            values = {}
            for offset, key in enumerate(config.METRIC_KEYS):
                cell = row[first - 1 + offset]
                values[key] = None if cell is None else float(cell)
            if all(v is not None for v in values.values()):
                slots[slot] = values
        if slots:
            out.append((when, slots))
    return out


def run_import(source: Path | None = None, db_path: Path | None = None,
               rebuild: bool = False) -> dict:
    """Import the workbook. With `rebuild`, wipe existing readings first."""
    import openpyxl  # lazy: see the note at the top of this module

    source = Path(source or config.SOURCE_XLSX)
    if not source.exists():
        raise FileNotFoundError(f"Workbook not found: {source}")

    db.init_db(db_path)
    # data_only: the sheet's date column is half literal dates and half =A3+1
    # formulas, and only the cached values are any use here.
    wb = openpyxl.load_workbook(source, data_only=True, read_only=True)
    try:
        rows = _read_rows(wb[SHEET])
    finally:
        wb.close()

    inserted = interpolated = 0
    with db.transaction(db_path) as conn:
        if rebuild:
            conn.execute("DELETE FROM readings")

        for when, slots in rows:
            # A day is treated as interpolated only if every weigh-in on it
            # looks interpolated - one odd value in an otherwise real pair is
            # far more likely to be a typo than a back-fill.
            estimated = all(_looks_interpolated(v) for v in slots.values())
            for slot, values in sorted(slots.items()):
                conn.execute(
                    f"""
                    INSERT INTO readings (day, slot, {mutations.READING_COLUMNS},
                                          estimated, note)
                    VALUES (?, ?, {mutations.READING_PLACEHOLDERS}, ?, ?)
                    ON CONFLICT (day, slot) DO UPDATE SET
                        {', '.join(f'{k} = excluded.{k}' for k in config.METRIC_KEYS)},
                        estimated  = excluded.estimated,
                        updated_at = datetime('now')
                    """,
                    (when.isoformat(), slot,
                     *[values[key] for key in config.METRIC_KEYS],
                     int(estimated), "Imported from the workbook"),
                )
                inserted += 1
            interpolated += bool(estimated)

        db.log(conn, "import", "workbook", None,
               f"{source.name}: {len(rows)} days, {inserted} readings")

    return {
        "days": len(rows),
        "readings": inserted,
        "detected_interpolated": interpolated,
        "first_day": rows[0][0].isoformat() if rows else None,
        "last_day": rows[-1][0].isoformat() if rows else None,
    }


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Import Weigh-in Tracker.xlsx into SQLite")
    parser.add_argument("--source", type=Path, default=config.SOURCE_XLSX)
    parser.add_argument("--db", type=Path, default=config.DB_PATH)
    parser.add_argument("--rebuild", action="store_true",
                        help="delete existing readings before importing")
    parser.add_argument("--no-backfill", action="store_true",
                        help="skip filling gaps left by the workbook")
    args = parser.parse_args()

    counts = run_import(args.source, args.db, rebuild=args.rebuild)
    print(f"Imported into {args.db}:")
    for label, value in counts.items():
        print(f"  {label:<24} {value}")

    if not args.no_backfill:
        # rebuild=False: fill days the sheet left blank, but leave the values
        # it back-filled itself alone, so the dashboard reconciles with the
        # spreadsheet rather than quietly recomputing six years of estimates.
        result = mutations.backfill_all(rebuild=False)
        print(f"\nBack-fill: {result['filled']} day(s) interpolated "
              f"between {result['anchors']} real readings")
        for line in result["skipped"]:
            print(f"  ! {line}")


if __name__ == "__main__":
    main()
