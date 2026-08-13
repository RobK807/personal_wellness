"""One-time (and re-runnable) import of the scraped Strava runs into SQLite.

Reads the Final_data tab of `strava_webscrape/strava_runs.xlsx` strictly
read-only. Run it again whenever the scrape is refreshed:

    python -m core.strava_import --rebuild

Sheet layout
------------
Final_data is one row per (run, breakdown), headings in row 1, data from row 2:

    A  Date              the run's date
    B  Distance (km)     the run
    C  Pace (min/km)     derived - not imported
    D  Time              the run's elapsed time
    E  Breakdown         which rung of the ladder this row is
    F  Breakdown time    the split
    G  Breakdown pace    derived - not imported
    H  Run type          Standard, Race, Weighted, Pace, Sprints, Intervals
    I  Effort type       Base, Threshold, Tempo, VO2 max, Race, Warm-up

Columns A-D and H-I repeat down every row of the same run: nine rows for a run
that reached nine rungs. That is a join flattened for a spreadsheet, and the
importer's whole job is to fold it back into `runs` and `run_bests`.

Neither pace column is imported. Both are quotients of two columns that are,
and importing a derived value is how a database ends up disagreeing with
itself. See the note in core/runs.py for the one-second difference this makes
to the breakdown pace.

Identifying a run
-----------------
Final_data has no activity id - the Clean_data tab has one, but the brief names
Final_data, and it is the tab that has been cleaned to be authoritative. Date,
distance and elapsed time together are what is left, and they are enough: the
1,546 rows fold into 229 runs, and the eleven dates carrying more than one run
(three on 07/06/2026) separate cleanly. It also means re-importing updates
rather than duplicates.

What is not corrected
---------------------
Four runs have `0` in the Breakdown column and no splits at all. They are
imported as runs with an empty ladder, which is what they are.

Eighteen splits used to be longer than the run they sat inside; the runs'
elapsed times were corrected in the sheet on 12/08/2026 and none is left. What
the importer does with any that come back is unchanged: they are imported as
they stand, because the alternative is a dashboard that quietly holds less than
the spreadsheet it was built from, and then counted here and listed on the
Admin page by core.run_queries.anomalies(). Nothing entered through the form
can add to them.

Forty-nine best efforts are still slower than their run's own average pace.
That is odd but, unlike the above, not strictly impossible - the fastest 5 km
inside a 5.4 km run really can be slower than the whole thing if the last 400 m
was quick - so it is reported rather than flagged.

Eleven breakdown *paces* used to be truncated: the cleaning took a fixed four
characters out of Strava's "5:16/km", one short of what "12:35/km" needs, so
every pace of ten minutes or more lost its last digit. `strava_runs!Q` and `!U`
now cut at the "/" instead. It never reached the database either way - neither
pace column is imported - and run_test.py asserts the column stays clean.
"""
from __future__ import annotations

import argparse
import datetime as dt
from pathlib import Path

import config
from core import db, run_mutations, runs

# openpyxl is imported lazily inside run_import(). Importing this module must
# stay cheap: the Flask front-end runs on a NAS with ~150 MB of RAM free, and
# the workbook is normally imported once on a desktop rather than on the NAS.

FIRST_ROW = 2
MAX_BLANK_RUN = 5  # stop scanning after this many consecutive dateless rows

# Column positions, 0-indexed into the tuple iter_rows hands back.
DATE, DISTANCE, _PACE, TIME, BREAKDOWN, SPLIT, _SPLIT_PACE, RUN_TYPE, EFFORT = range(9)

# Excel keeps a date as a count of days and a duration as a fraction of one,
# and it is the cell's *number format* that decides whether openpyxl hands
# back a date or the bare number. Final_data is a paste-values sheet, so that
# formatting is one careless paste away from being lost - and it has been, on
# the oldest 51 rows.
#
# Losing a number format does not lose the value, so the importer converts
# rather than refusing: a run that reads 43744 is still 06/10/2019. The epoch
# is the 30th of December rather than the 31st because Excel believes 1900 was
# a leap year and openpyxl reproduces that.
EXCEL_EPOCH = dt.date(1899, 12, 30)

# Serials outside this range are not plausible dates for a running log, and
# accepting them would turn a stray number into a run in 1901. The bounds are
# 01/01/1900 and 01/01/2064.
EXCEL_DATE_RANGE = (1, 60000)

# What a failed formula leaves behind in a pasted cell. Clean_data looks the
# run and effort types up with OFFSET/MATCH over a fixed range, so a run whose
# id is not in that range comes through as the text "#N/A".
#
# Imported as UNCLASSIFIED rather than verbatim: the run happened and its
# distance, time and splits are all fine, so throwing it away would lose real
# data - but "#N/A" is not a kind of run, and letting it through would put it
# in the analysis as though it were one. The count is reported so it gets
# fixed at the source.
EXCEL_ERRORS = frozenset(("#N/A", "#REF!", "#VALUE!", "#DIV/0!", "#NAME?",
                          "#NULL!", "#NUM!", "#SPILL!", "#CALC!"))
UNCLASSIFIED = "Unclassified"


def _classification(value, fallback: str) -> tuple:
    """A run or effort type, and whether the sheet actually knew it."""
    text = str(value or "").strip()
    if text in EXCEL_ERRORS:
        return UNCLASSIFIED, False
    return (text or fallback), True


def _excel_date(value, where: str) -> dt.date:
    """A date from a formatted cell, or from a bare Excel day count."""
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        low, high = EXCEL_DATE_RANGE
        if not low <= value <= high:
            raise ValueError(
                f"{where}: {value!r} is not a plausible Excel date serial "
                f"(expected between {low} and {high})")
        return EXCEL_EPOCH + dt.timedelta(days=int(value))
    return runs.as_date(value)


def _excel_seconds(value, where: str) -> int:
    """Seconds from a formatted time cell, or from a bare fraction of a day.

    A number here is always days, never seconds - this is an Excel cell, and
    that is what Excel stores. core.runs.parse_duration() reads a bare number
    as seconds, which is right for something a person typed into a form and
    wrong for something read out of a spreadsheet, so the two do not share a
    rule.
    """
    if isinstance(value, (dt.time, dt.timedelta, str)):
        return runs.parse_duration(value, where)
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return int(round(float(value) * 86400))
    return runs.parse_duration(value, where)


def _read_rows(ws) -> list:
    """Fold the sheet into [(identity, run fields, [ladder rows])].

    Iterated rather than addressed cell by cell: openpyxl's read-only mode
    streams the sheet, so `ws.cell(r, c)` rescans it from the top every call.

    Rows for the same run are not assumed to be adjacent - they are in this
    sheet, but grouping on the identity rather than on "the date changed" costs
    nothing and does not care.
    """
    grouped: dict = {}
    order: list = []
    blanks = 0
    unformatted = 0

    for number, row in enumerate(ws.iter_rows(min_row=FIRST_ROW, max_col=9,
                                              values_only=True),
                                 start=FIRST_ROW):
        if row[DATE] is None:
            blanks += 1
            if blanks > MAX_BLANK_RUN:
                break
            continue
        blanks = 0

        if not isinstance(row[DATE], (dt.date, dt.datetime)):
            unformatted += 1
        when = _excel_date(row[DATE], f"Row {number}, date")
        distance = round(float(row[DISTANCE]), 2)
        duration = _excel_seconds(row[TIME], f"Row {number}, time")
        identity = (when, distance, duration)

        if identity not in grouped:
            order.append(identity)
            run_type, known_run = _classification(row[RUN_TYPE], "Standard")
            effort_type, known_effort = _classification(row[EFFORT], "Base")
            grouped[identity] = {
                "day": when,
                "distance_km": distance,
                "duration_s": duration,
                "run_type": run_type,
                "effort_type": effort_type,
                "classified": known_run and known_effort,
                "ladder": [],
            }

        label = row[BREAKDOWN]
        # `0` is the sheet's way of saying "no breakdown for this run" - four
        # rows have it, and they are the run rather than a split of it.
        if label is None or label == 0 or str(label).strip() in ("", "0"):
            continue

        label = str(label).strip()
        if label not in config.BREAKDOWN_KM:
            raise ValueError(
                f"Row for {when:%d/%m/%Y}: '{label}' is not a known breakdown "
                f"distance. Add it to config.BREAKDOWNS or fix the sheet.")

        grouped[identity]["ladder"].append({
            "breakdown": label,
            "ordinal": config.BREAKDOWN_ORDER[label],
            "km": config.BREAKDOWN_KM[label],
            "seconds": _excel_seconds(row[SPLIT], f"Row {number}, {label} split"),
        })

    for entry in grouped.values():
        entry["ladder"].sort(key=lambda item: item["ordinal"])
    return [grouped[identity] for identity in order], unformatted


def _count_anomalies(entries: list) -> dict:
    """What the sheet contains that cannot physically be true. See the header."""
    longer_than_run = 0
    out_of_order = 0
    too_far = 0
    for entry in entries:
        for row in entry["ladder"]:
            if row["seconds"] > entry["duration_s"]:
                longer_than_run += 1
            if row["km"] > entry["distance_km"] * 1.001:
                too_far += 1
        for shorter, longer in zip(entry["ladder"], entry["ladder"][1:]):
            if longer["seconds"] < shorter["seconds"]:
                out_of_order += 1
    return {"splits_longer_than_run": longer_than_run,
            "splits_out_of_order": out_of_order,
            "splits_longer_than_distance": too_far}


def run_import(source: Path | None = None, sheet: str | None = None,
               db_path: Path | None = None, rebuild: bool = False) -> dict:
    """Import the run workbook. With `rebuild`, wipe existing runs first.

    Without it, runs are matched on (date, distance, time) and updated in
    place, so a refreshed scrape brings the new runs across and corrects any
    reclassified ones without disturbing runs entered by hand.
    """
    import openpyxl  # lazy: see the note at the top of this module

    source = Path(source or config.RUNS_XLSX)
    sheet = sheet or config.RUNS_SHEET
    if not source.exists():
        raise FileNotFoundError(f"Run workbook not found: {source}")

    db.init_db(db_path)
    wb = openpyxl.load_workbook(source, data_only=True, read_only=True)
    try:
        if sheet not in wb.sheetnames:
            raise KeyError(f"{source.name} has no '{sheet}' tab "
                           f"(it has {', '.join(wb.sheetnames)})")
        entries, unformatted = _read_rows(wb[sheet])
    finally:
        wb.close()

    splits = 0
    restored = orphaned = 0
    with db.transaction(db_path) as conn:
        # Interval details are entered by hand and are nowhere in the sheet, so
        # a --rebuild would otherwise throw them away along with the rows they
        # hang off. Kept against the run's identity and put back afterwards.
        kept = {
            (row["day"], row["distance_km"], row["duration_s"]): dict(row)
            for row in conn.execute(
                "SELECT day, distance_km, duration_s, interval_type, "
                "interval_count, interval_distance_m, interval_time_s, "
                "interval_pace_s "
                "FROM runs WHERE interval_type IS NOT NULL")
        }

        if rebuild:
            # run_bests goes with it through ON DELETE CASCADE.
            conn.execute("DELETE FROM runs WHERE source = 'strava'")

        for entry in entries:
            run_mutations.load_run(
                conn,
                day=entry["day"],
                distance_km=entry["distance_km"],
                duration_s=entry["duration_s"],
                run_type=entry["run_type"],
                effort_type=entry["effort_type"],
                ladder=entry["ladder"],
                source="strava",
            )
            splits += len(entry["ladder"])

        for identity, saved in kept.items():
            changed = conn.execute(
                """
                UPDATE runs SET interval_type = ?, interval_count = ?,
                                interval_distance_m = ?, interval_time_s = ?,
                                interval_pace_s = ?
                WHERE day = ? AND distance_km = ? AND duration_s = ?
                  AND interval_type IS NULL
                """,
                (saved["interval_type"], saved["interval_count"],
                 saved["interval_distance_m"], saved["interval_time_s"],
                 saved["interval_pace_s"], *identity),
            ).rowcount
            restored += changed
            # The run's date, distance or time changed in the sheet, so there is
            # nothing left with the identity the session was recorded against.
            orphaned += not changed and identity not in {
                (e["day"].isoformat(), e["distance_km"], e["duration_s"])
                for e in entries}

        db.log(conn, "import", "runs", None,
               f"{source.name}[{sheet}]: {len(entries)} runs, {splits} splits")

    days = sorted(entry["day"] for entry in entries)
    return {
        "runs": len(entries),
        "splits": splits,
        "without_splits": sum(1 for e in entries if not e["ladder"]),
        "rows_missing_a_date_format": unformatted,
        "runs_the_sheet_could_not_classify":
            sum(1 for e in entries if not e["classified"]),
        "interval_sessions_kept": restored,
        "interval_sessions_orphaned": orphaned,
        "first_day": days[0].isoformat() if days else None,
        "last_day": days[-1].isoformat() if days else None,
        "distance_km": round(sum(e["distance_km"] for e in entries), 2),
        "duration_s": sum(e["duration_s"] for e in entries),
        **_count_anomalies(entries),
    }


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Import the scraped Strava runs into SQLite")
    parser.add_argument("--source", type=Path, default=config.RUNS_XLSX)
    parser.add_argument("--sheet", default=config.RUNS_SHEET)
    parser.add_argument("--db", type=Path, default=config.DB_PATH)
    parser.add_argument("--rebuild", action="store_true",
                        help="delete previously imported runs before importing")
    args = parser.parse_args()

    counts = run_import(args.source, args.sheet, args.db, rebuild=args.rebuild)
    print(f"Imported into {args.db}:")
    for label, value in counts.items():
        if label == "duration_s":
            value = runs.fmt_duration(value, force_hours=True)
        print(f"  {label:<28} {value}")

    if counts["interval_sessions_orphaned"]:
        print(f"\n{counts['interval_sessions_orphaned']} hand-entered interval "
              f"session(s) could not be put back: the run's date, distance or "
              f"time has changed in the sheet, so there is nothing left with "
              f"the identity they were recorded against. Re-enter them on the "
              f"run's page.")

    if counts["runs_the_sheet_could_not_classify"]:
        print(f"\n{counts['runs_the_sheet_could_not_classify']} run(s) came "
              f"through with #N/A for their run or effort type and have been "
              f"imported as '{UNCLASSIFIED}'. Clean_data looks both up with "
              f"OFFSET/MATCH over a fixed range; a run whose id falls outside "
              f"it will not resolve.")

    if counts["rows_missing_a_date_format"]:
        print(f"\n{counts['rows_missing_a_date_format']} rows have lost their "
              "date and time number formats in the sheet and were read as raw "
              "Excel serials. The values are intact, so the import is correct; "
              "reformatting those cells as Date and Time would restore the tab "
              "to being readable.")

    if counts["splits_longer_than_run"] or counts["splits_out_of_order"]:
        print("\nThe sheet contains splits that cannot be true - see the note "
              "at the top of core/strava_import.py. They have been imported as "
              "they stand and are listed on the run tracker's Admin page.")


if __name__ == "__main__":
    main()
