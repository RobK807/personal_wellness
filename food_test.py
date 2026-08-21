"""Check the food section against the workbook and the corrected diary CSV.

    python food_test.py

The run tracker has run_test.py and the workouts have workout_test.py; this is
the same idea for the food section. Read `Food Planner v0.1.xlsx` here, count
what is in it, load the corrected CSV, and assert the database holds exactly
that.

Six things are checked.

**The catalogue.** The Food sheet's three lists - Items at B, Meals at N,
Recipes at Z - become 183 rows, plus the four in config.FOOD_EXTRA_FOODS that
the corrected diary refers to and the sheet never had.

**The diary.** Every entry in the 49 week blocks reaches the database with its
macros intact, and the days the workbook itself totals reproduce to the penny.
The workbook's own Consumed row is the reference, not a figure of mine.

**The portion arithmetic.** A catalogue row records its macros *for* a portion,
so eating half of one is half of it. This is the only arithmetic in the section
that can silently be wrong, so it is checked against foods whose portions are
not 1 - the rice at 75 g and the pasta at 90 g.

**The week.** Seven days from whichever day the week starts on, with the days
that have nothing against them coming back as gaps rather than as zeros.

**The calculator.** Components added up and then scaled, with a catalogue
component scaled by its own portion first and a typed one multiplied by its
quantity.

**The write path.** A day saved through the mutations reads back the way it was
written, a day is replaced rather than appended to, and the validation refuses
what it should.

It runs against a throwaway copy, so the real database is untouched.
"""
from __future__ import annotations

import datetime as dt
import os
import sys
import tempfile
from pathlib import Path

TEMP_DB = Path(tempfile.gettempdir()) / "wellness_food_test.db"
for suffix in ("", "-wal", "-shm"):
    Path(str(TEMP_DB) + suffix).unlink(missing_ok=True)
os.environ["PW_DB_PATH"] = str(TEMP_DB)

import config  # noqa: E402
from core import (db, food, food_import, food_mutations as fm,  # noqa: E402
                  food_queries as fq)

# The corrected export, hand-checked and handed back. The raw one is what
# --export writes; this is what --load reads.
DIARY = Path(__file__).parent / "data" / "imports" / "food_diary_cleaned.csv"

failures: list = []


def check(label: str, actual, expected) -> None:
    ok = actual == expected
    shown = repr(actual)
    print(f"  [{'ok ' if ok else 'FAIL'}] {label}: "
          f"{shown[:150]}{'...' if len(shown) > 150 else ''}"
          + ("" if ok else f" (expected {expected!r})"))
    if not ok:
        failures.append(label)


def close(label: str, actual, expected, tolerance: float = 0.01) -> None:
    ok = actual is not None and abs(float(actual) - float(expected)) <= tolerance
    print(f"  [{'ok ' if ok else 'FAIL'}] {label}: {actual}"
          + ("" if ok else f" (expected {expected})"))
    if not ok:
        failures.append(label)


# --------------------------------------------------------------------------- #
# The workbook, read here rather than through the importer
# --------------------------------------------------------------------------- #
def read_workbook() -> dict:
    """Count the sheet independently, and pull out the days it totals itself."""
    import openpyxl

    wb = openpyxl.load_workbook(config.FOOD_XLSX, data_only=True)
    try:
        food_sheet = wb["Food"]
        catalogue = 0
        for column in food_import.CATALOGUE_BLOCKS.values():
            for row in range(food_import.CATALOGUE_FIRST_ROW,
                             food_sheet.max_row + 1):
                name = food_sheet.cell(row, column).value
                macros = [food_sheet.cell(row, column + offset).value
                          for offset in (4, 5, 6, 7)]
                if name and all(value is not None for value in macros):
                    catalogue += 1

        # The diary's own totals. A day block is: the weekday name, a header
        # row, sixteen entry rows, then "Target" and "Consumed". Consumed is
        # the sheet's own arithmetic rather than mine, which is what makes it
        # worth comparing against - and it is the row *below* Target, which is
        # where a 1,890 that is the target rather than the total comes from.
        #
        # One record per *block*, not per day, and each carries both the
        # Consumed cell and the sum of the sixteen rows above it. Two blocks
        # claim w/c 27/07/2026, so a dict keyed by day would silently keep
        # whichever came last; and comparing a block's total against its own
        # rows is what "the sheet contradicts itself" actually means.
        diary = wb["Food_Diary"]
        consumed_offset = 3 + food_import.ENTRIES_PER_DAY
        blocks = []
        for column in range(1, diary.max_column + 1):
            header = diary.cell(1, column).value
            match = food_import.WEEK_HEADER.search(str(header or ""))
            if not match:
                continue
            monday = dt.date(int(match.group(3)), int(match.group(2)),
                             int(match.group(1)))
            for index in range(7):
                start = food_import.DAY_FIRST_ROW + food_import.DAY_STRIDE * index
                label = diary.cell(start + consumed_offset, column + 1).value
                if str(label or "").strip().casefold() != "consumed":
                    continue      # not the shape this expects; say nothing
                value = diary.cell(start + consumed_offset, column + 2).value
                if not isinstance(value, (int, float)) or not value:
                    continue
                rows = sum(
                    cell for cell in
                    (diary.cell(row, column + 2).value
                     for row in range(start + 2,
                                      start + 2 + food_import.ENTRIES_PER_DAY))
                    if isinstance(cell, (int, float)))
                blocks.append({"day": monday + dt.timedelta(days=index),
                               "consumed": float(value), "rows": float(rows)})
    finally:
        wb.close()
    return {"catalogue": catalogue, "blocks": blocks}


def _workbook_entry_totals() -> dict:
    """Calories per day from the diary's entry rows - the sheet's raw record.

    Read through the importer's own fold rather than a second reading of the
    sheet, because the point of comparing against this is to find what the CSV
    changed, not to re-test read_diary.
    """
    totals: dict = {}
    for entry in food_import.read_diary(Path(config.FOOD_XLSX))["entries"]:
        totals[entry["date"]] = totals.get(entry["date"], 0.0) + entry["calories"]
    return totals


def _csv_day_totals() -> dict:
    """Calories per day straight out of the corrected CSV.

    Its own reading of the file, deliberately: comparing the database against
    the loader's report of what it loaded would only prove the loader agrees
    with itself.
    """
    import csv

    totals: dict = {}
    with DIARY.open(encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            day = food_import._csv_date(row)
            totals[day] = totals.get(day, 0.0) + float(row["calories"])
    return totals


# --------------------------------------------------------------------------- #
def main() -> int:
    if not Path(config.FOOD_XLSX).exists():
        print(f"No workbook at {config.FOOD_XLSX} - nothing to reconcile.")
        return 1
    if not DIARY.exists():
        print(f"No corrected diary at {DIARY} - nothing to reconcile.")
        return 1

    print(f"Workbook: {config.FOOD_XLSX}")
    print(f"Diary   : {DIARY}")
    print(f"Test db : {TEMP_DB}\n")

    sheet = read_workbook()
    db.init_db()

    # --- the catalogue --------------------------------------------------------
    print("The catalogue")
    loaded = food_import.import_catalogue(config.FOOD_XLSX)
    extras = len(config.FOOD_EXTRA_FOODS)
    check("rows read from the sheet, plus the configured extras",
          loaded["read"], sheet["catalogue"] + extras)
    check("all of them added to an empty database",
          loaded["added"], sheet["catalogue"] + extras)
    check("the three lists are all present",
          sorted(loaded["by_list"]), sorted(config.FOOD_LISTS))
    check("re-importing changes nothing",
          food_import.import_catalogue(config.FOOD_XLSX)["added"], 0)

    # The workbook has "Mashed potato" as both an Item and a Recipe, which is
    # why the catalogue is unique on (list, name) rather than on name.
    duplicated = fq.foods(search="Mashed potato", include_retired=True)
    check("the same name in two lists is two rows", len(duplicated), 2)
    check("in different lists",
          sorted(row["list"] for row in duplicated), ["Items", "Recipes"])

    # --- the diary ------------------------------------------------------------
    print("\nThe diary")
    result = food_import.load_diary(DIARY, replace=True)
    check("days", result["days"], 290)
    check("entries", result["entries"], 2065)
    check("every line names something in the catalogue",
          result["free_text"], 0)
    coverage = fq.coverage()
    check("first day", coverage["first_day"], "2024-01-02")
    check("last day", coverage["last_day"], "2026-08-08")

    # --- the diary reconciles, in three steps ---------------------------------
    #
    # Three comparisons rather than one, because three different things can be
    # true at once and lumping them together hides all of them:
    #
    #   the database vs the CSV        did the load do what the file said
    #   the CSV vs the workbook        what was corrected by hand, deliberately
    #   the workbook vs itself         where the sheet contradicts its own sums
    #
    print("\nThe database holds exactly what the CSV said")
    from_csv = _csv_day_totals()
    differing = [day for day, expected in from_csv.items()
                 if fq.day(day) is None
                 or abs(fq.day(day)["calories"] - expected) > 0.02]
    check("days in the file", len(from_csv), 290)
    check("days where the database and the file disagree", differing, [])

    print("\nWhat the corrected CSV changed, on purpose")
    entry_totals = _workbook_entry_totals()
    corrected = sorted(day for day, expected in from_csv.items()
                       if abs(expected - entry_totals.get(day, 0)) > 0.02)
    # Corrections by hand, and the count is pinned rather than waved at: the
    # three rows the export flagged, the dinners-out re-estimated onto one
    # standing figure, a handful of recipes refreshed, and a "Pizza Night" that
    # the workbook recorded as zero calories.
    check("days the CSV deliberately restates", len(corrected), 17)
    check("and every other day matches the workbook line for line",
          len(from_csv) - len(corrected), 273)

    print("\nWhere the workbook disagrees with itself")
    # A day block is sixteen entry rows and then a "Consumed" row that sums
    # them. Six times it does not: the cached formula value is stale, so the
    # sheet's own total contradicts the entries printed directly above it. The
    # entries are the record and the total is a formula, so the entries win -
    # but the disagreement is pinned here rather than smoothed over.
    stale = sorted({block["day"] for block in sheet["blocks"]
                    if abs(block["consumed"] - block["rows"]) > 0.05})
    check("blocks whose Consumed cell contradicts its own entry rows",
          [day.isoformat() for day in stale],
          ["2024-01-12", "2024-01-13", "2024-01-14", "2024-12-02",
           "2026-08-01", "2026-08-02"])

    print("\nEvery other day matches the workbook's own arithmetic")
    # The sheet's Consumed row, on the days where the sheet agrees with itself
    # and the CSV did not restate anything. This is the real reconciliation:
    # the workbook's arithmetic, not mine. A day claimed by two blocks passes
    # if it matches either, because which of the two was kept is read_diary's
    # decision and is asserted where that decision is made.
    skip = set(stale) | set(corrected)
    by_day: dict = {}
    for block in sheet["blocks"]:
        by_day.setdefault(block["day"], []).append(block["consumed"])

    compared, wrong = 0, []
    for day, values in sorted(by_day.items()):
        row = fq.day(day)
        if day in skip or row is None:
            continue
        compared += 1
        if all(abs(row["calories"] - value) > 0.02 for value in values):
            wrong.append((day.isoformat(), row["calories"], values))
    check("days compared against the sheet's own Consumed row",
          compared >= 250, True)
    check("days that differ", wrong, [])

    # One day spelled out, so a failure above has something to be read against.
    close("03/08/2026 reproduces the workbook exactly",
          fq.day("2026-08-03")["calories"], 2036.3)

    # --- portions -------------------------------------------------------------
    print("\nPortion arithmetic, on foods whose portion is not 1")
    rice = fq.food_by_name("Sainsbury's basmati rice")
    check("the rice is recorded per 75 g", rice["portion"], 75.0)
    close("75 g is one portion", food.eaten(rice, 75)["calories"],
          rice["calories"])
    close("150 g is two", food.eaten(rice, 150)["calories"],
          rice["calories"] * 2)
    close("37.5 g is half", food.eaten(rice, 37.5)["calories"],
          rice["calories"] / 2)
    close("and the carbs scale with it", food.eaten(rice, 150)["carbs"],
          rice["carbs"] * 2)

    pasta = fq.food_by_name("Sainsbury's wheat pasta")
    check("the pasta is recorded per 90 g", pasta["portion"], 90.0)
    close("50 g of it", food.eaten(pasta, 50)["calories"],
          pasta["calories"] * 50 / 90)

    # A half-filled catalogue row is shown at face value rather than refused:
    # showing nothing is worse than showing what is there.
    check("a zero portion is treated as one",
          food.portion_factor({"portion": 0}, 3), 3.0)

    # --- weeks ----------------------------------------------------------------
    print("\nWeeks")
    monday = dt.date(2026, 8, 3)
    week = fq.week(monday, 0)
    check("a Monday week starts on the Monday", week["start"], monday)
    check("seven days", len(week["days"]), 7)
    check("labelled the way the workbook labels them", week["label"],
          "w/c 03/08/2026")
    # The workbook recorded six of these seven days; the seventh is a gap and
    # must come back as one rather than as a recorded zero.
    check("six planned, one blank",
          sum(1 for row in week["days"] if row["planned"]), 6)
    close("the week totals its planned days", week["totals"]["calories"],
          sum(row["calories"] for row in week["days"]))

    thursday = fq.week(monday, 3)
    check("a Thursday week starts on the Thursday",
          thursday["start"], dt.date(2026, 7, 30))
    check("and still holds seven days", len(thursday["days"]), 7)
    check("every start day gives a week containing the anchor",
          all(fq.week(monday, start)["start"] <= monday
              <= fq.week(monday, start)["end"] for start in range(7)), True)

    # --- the calculator -------------------------------------------------------
    print("\nThe macro calculator")
    result = fq.calculate([
        {"food_id": rice["id"], "quantity": 150},
        {"name": "Off a packet", "quantity": 200, "calories": 2, "carbs": 1,
         "fat": 0.5, "protein": 0.25},
        {"name": "", "quantity": "", "calories": "", "carbs": "", "fat": "",
         "protein": ""},
    ], scale=0.5)
    check("the blank row is not a component", len(result["components"]), 2)
    close("the catalogue row is scaled by its own portion",
          result["components"][0]["calories"], rice["calories"] * 2)
    close("a typed row is multiplied by its quantity",
          result["components"][1]["calories"], 400.0)
    close("the subtotal adds them", result["subtotal"]["calories"],
          rice["calories"] * 2 + 400.0)
    close("and the scale applies to the total, not to a row",
          result["total"]["calories"], (rice["calories"] * 2 + 400.0) / 2)
    close("every macro is scaled, not only the calories",
          result["total"]["protein"],
          (rice["protein"] * 2 + 50.0) / 2)

    # --- splitting the historic strings ---------------------------------------
    print("\nReading the workbook's rendered diary lines")
    cases = [
        ("Banana - 1 Portion", "Banana", 1.0, "Portion"),
        # A name that contains its own " - " separator. The split is anchored
        # on the last " - number units", not the first separator, or this one
        # becomes "Nature Valley" eaten in units of "Salted Caramel Nut".
        ("Nature Valley - Salted Caramel Nut - 1 Bar",
         "Nature Valley - Salted Caramel Nut", 1.0, "Bar"),
        ("Sainsbury's basmati rice - 85 grams",
         "Sainsbury's basmati rice", 85.0, "grams"),
        # Nothing to split: the honest answer is the whole string as the name.
        ("Dinner out", "Dinner out", None, None),
    ]
    for text, name, quantity, units in cases:
        parts = food.split_description(text)
        check(f"{text!r}", (parts["name"], parts["quantity"], parts["units"]),
              (name, quantity, units))

    # --- targets --------------------------------------------------------------
    print("\nTargets are dated, so a day is measured against its own time")
    fm.save_target({"name": "Test", "starts_on": "2025-01-01",
                    "calories": 2000, "carbs": 200, "fat": 60, "protein": 150})
    fm.save_target({"name": "Test", "starts_on": "2026-01-01",
                    "calories": 1800, "carbs": 180, "fat": 50, "protein": 180})
    check("a 2025 day reads the 2025 version",
          fq.target_for("2025-06-01", "Test")["calories"], 2000.0)
    check("a 2026 day reads the 2026 one",
          fq.target_for("2026-06-01", "Test")["calories"], 1800.0)
    check("a day before either reads neither",
          fq.target_for("2024-06-01", "Test"), None)

    # --- the write path -------------------------------------------------------
    print("\nThe write path")
    when = dt.date(2027, 1, 15)
    fm.save_day(when, [
        {"meal": "Breakfast", "food_id": rice["id"], "quantity": 150},
        {"meal": "Dinner", "name": "Dinner out", "quantity": 1,
         "units": "Portion", "calories": 800, "carbs": 70, "fat": 30,
         "protein": 40},
    ])
    row = fq.day(when)
    check("two lines", row["entries"], 2)
    close("the catalogue line was scaled on the way in",
          row["calories"], rice["calories"] * 2 + 800)
    check("the linked line keeps its food",
          fq.entries(when)[0]["food_id"], rice["id"])

    fm.save_day(when, [{"meal": "Lunch", "name": "Only this", "quantity": 1,
                        "calories": 100, "carbs": 10, "fat": 1, "protein": 5}])
    check("saving a day replaces it rather than adding to it",
          fq.day(when)["entries"], 1)

    fm.copy_day(when, when + dt.timedelta(days=1))
    close("a copied day carries the same macros",
          fq.day(when + dt.timedelta(days=1))["calories"], 100.0)

    made = fm.fill_week(when, when, starts_on=0, overwrite=False)
    check("filling a week skips the days that already have entries",
          [day.isoformat() for day in made["skipped"]],
          [(when + dt.timedelta(days=1)).isoformat()])
    check("and fills the rest", len(made["copied"]), 5)

    print("\nValidation refuses what it should")
    for label, values in [
        ("a day of 99,999 calories", {"meal": "Lunch", "name": "x",
                                      "calories": 99999, "carbs": 0, "fat": 0,
                                      "protein": 0}),
        ("a line with no food and no name", {"meal": "Lunch", "name": ""}),
        ("a negative macro", {"meal": "Lunch", "name": "x", "calories": -5,
                              "carbs": 0, "fat": 0, "protein": 0}),
    ]:
        try:
            food.parse_entry(values)
        except food.InvalidFood:
            check(label, "refused", "refused")
        else:
            check(label, "accepted", "refused")

    try:
        fm.save_food({"name": "Rubbish", "list": "Nonsense", "portion": 1,
                      "calories": 1, "carbs": 0, "fat": 0, "protein": 0})
    except food.InvalidFood:
        check("a food in a list that does not exist", "refused", "refused")
    else:
        check("a food in a list that does not exist", "accepted", "refused")

    print()
    if failures:
        print(f"FAILED: {len(failures)} check(s): {', '.join(failures)}")
        return 1
    print("The workbook reconciles: every catalogue row, every diary line, "
          "and every day the sheet totals for itself.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
