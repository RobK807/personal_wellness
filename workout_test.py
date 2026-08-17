"""Check the workout tracker against the gym workbook it was built from.

    python workout_test.py

The run tracker has run_test.py, which re-folds the Strava sheet independently
and asserts every deviation. This is the same idea for the workout section: read
`2026 Gym Programme.xlsx` here, count what is in it, and assert the database
holds exactly that.

Four things are checked.

**The fold.** 19 week sheets become 19 weeks, 38 sessions and 564 sets, with the
right phase, cycle type and coaching notes on each week.

**The weights.** Every one of the workbook's Weight (kg) cells that sits beside a
percentage is `% x 1RM` rounded to 2.5 kg. None of them is imported - the
percentage is - so this recomputes all of them from the sheet's own 1RMs and
compares. If the sheet is ever edited without recalculating, this is what says
so.

**The rounding rule.** SQLite's ROUND is half-away-from-zero and Python's round()
is half-to-even, so 61.25 kg at a 2.5 step is 62.5 to one and 60.0 to the other.
The database computes prescribed weights in v_exercise_sets and the input form
previews them with core.workouts.round_to(); this asserts the two never disagree,
including on the exact half-way cases where they would.

**The write path.** A plan built through the mutations reads back the way it was
written, a copied plan reproduces the structure without the tick-offs, and the
validation refuses the things it should.

It runs against a throwaway copy, so the real database is untouched.
"""
from __future__ import annotations

import os
import sys
import tempfile
from pathlib import Path

TEMP_DB = Path(tempfile.gettempdir()) / "wellness_workout_test.db"
for suffix in ("", "-wal", "-shm"):
    Path(str(TEMP_DB) + suffix).unlink(missing_ok=True)
os.environ["PW_DB_PATH"] = str(TEMP_DB)

import config  # noqa: E402
from core import (db, gym_import, workout_mutations as wm,  # noqa: E402
                  workout_queries as wq, workouts)

failures: list = []


def check(label: str, actual, expected) -> None:
    ok = actual == expected
    shown = repr(actual)
    print(f"  [{'ok ' if ok else 'FAIL'}] {label}: "
          f"{shown[:150]}{'...' if len(shown) > 150 else ''}"
          + ("" if ok else f" (expected {expected!r})"))
    if not ok:
        failures.append(label)


# --------------------------------------------------------------------------- #
# The workbook, read here rather than through the importer
# --------------------------------------------------------------------------- #
def read_workbook() -> dict:
    """Count the sheet independently: weeks, sessions, exercises, sets."""
    import openpyxl

    wb = openpyxl.load_workbook(config.GYM_XLSX, data_only=True)
    weeks, sets, sessions, exercises = 0, 0, 0, 0
    percent_weights = []
    maxes = {}
    try:
        overview = wb["Programme Overview"]
        for row in range(5, 9):
            name = overview.cell(row, 1).value
            value = overview.cell(row, 2).value
            if name and isinstance(value, (int, float)):
                maxes[str(name).strip()] = float(value)

        for ws in wb.worksheets:
            if not ws.title.lower().startswith("week"):
                continue
            weeks += 1
            current = None
            for values in ws.iter_rows(min_row=4, max_col=8, values_only=True):
                first = str(values[0] or "").strip()
                kind = str(values[1] or "").strip().lower()
                if kind in ("warm-up", "working", "accessory"):
                    sets += 1
                    if isinstance(values[4], (int, float)) and \
                            isinstance(values[5], (int, float)):
                        percent_weights.append(
                            (current, float(values[4]), float(values[5])))
                elif first.upper().startswith("SESSION"):
                    sessions += 1
                    current = None
                elif first and not first.startswith(("•", "ℹ", "*")) \
                        and "NOTES" not in first.upper() \
                        and "GUIDANCE" not in first.upper():
                    if first != current:
                        exercises += 1
                        current = first
    finally:
        wb.close()
    return {"weeks": weeks, "sessions": sessions, "exercises": exercises,
            "sets": sets, "percent_weights": percent_weights, "maxes": maxes}


def main() -> int:
    if not config.GYM_XLSX.exists():
        print(f"Gym workbook not found: {config.GYM_XLSX}")
        return 1

    print(f"Reading {config.GYM_XLSX.name}")
    sheet = read_workbook()
    print(f"  {sheet['weeks']} week sheets, {sheet['sessions']} sessions, "
          f"{sheet['exercises']} exercises, {sheet['sets']} sets")

    result = gym_import.run_import(rebuild=True)
    plan = wq.plan(result["plan_id"])

    # --- the fold ----------------------------------------------------------
    print("\nThe fold")
    check("weeks imported", result["weeks"], sheet["weeks"])
    check("sessions imported", result["sessions"], sheet["sessions"])
    check("exercises imported", result["exercises"], sheet["exercises"])
    check("sets imported", result["sets"], sheet["sets"])
    check("weeks in the database", plan["weeks"], sheet["weeks"])
    check("sessions in the database", plan["sessions"], sheet["sessions"])
    check("sets in the database",
          db.scalar("SELECT COUNT(*) FROM exercise_sets", default=0),
          sheet["sets"])
    check("1RMs", result["one_rms"], sheet["maxes"])
    check("no exercise was unknown to the catalogue",
          result["exercises_added"], [])

    weeks = wq.weeks(plan["id"])
    check("weeks are numbered 1..19", [row["number"] for row in weeks],
          list(range(1, 20)))
    check("the deload week is labelled",
          [row["number"] for row in weeks if row["label"]], [19])
    check("the A/B rotation alternates",
          "".join(row["cycle_type"] or "-" for row in weeks),
          "ABABABABABABABABAB-")
    check("every week but the deload has a phase",
          [row["number"] for row in weeks if not row["phase_name"]], [])
    check("every week carries its coaching notes",
          [row["number"] for row in weeks if not row["note"]], [])
    check("phases", [row["name"] for row in wq.phases(plan["id"])],
          ["Phase 1", "Phase 2", "Phase 3", "Deload"])

    # --- the weights -------------------------------------------------------
    print("\nWeights the sheet computed, recomputed here")
    # Asserted exactly rather than loosely. If the sheet gains or loses a
    # percentage set this should say so, not shrug.
    check("percentage-and-weight pairs in the sheet",
          len(sheet["percent_weights"]), 268)
    wrong = []
    for lift, percent, weight in sheet["percent_weights"]:
        one_rm = sheet["maxes"].get(lift)
        if one_rm is None:
            continue
        mine = workouts.round_to(percent * one_rm, 2.5)
        if abs(mine - weight) > 0.001:
            wrong.append((lift, percent, weight, mine))
    check("every one is % x 1RM rounded to 2.5 kg", wrong, [])
    check("and the importer agrees", result["weight_mismatches"], [])

    print("\nThe database works the same weights out")
    rows = db.rows(
        "SELECT exercise_name, percent_1rm, one_rm_kg, rounding_kg, "
        "prescribed_kg FROM v_exercise_sets WHERE plan_id = ? "
        "AND load_mode = 'percent'", (plan["id"],))
    check("the database holds the same number",
          len(rows), len(sheet["percent_weights"]))
    disagreed = [row for row in rows
                 if row["one_rm_kg"] is not None
                 and abs(workouts.round_to(row["percent_1rm"] * row["one_rm_kg"],
                                           row["rounding_kg"])
                         - row["prescribed_kg"]) > 1e-9]
    check("SQL and Python round identically", disagreed, [])
    check("no percentage set is left without a 1RM",
          wq.percent_sets_without_a_max(plan["id"]), [])

    # The half-way cases, which is where the two rules would part company. Done
    # against the database rather than in the abstract: it is v_exercise_sets
    # that has to agree, not a function that looks like it.
    print("\nHalf-way cases round the same way in both")
    ties = []
    for percent, one_rm, step, want in ((0.5, 122.5, 2.5, 62.5),
                                        (0.5, 127.5, 2.5, 65.0),
                                        (0.25, 10.0, 1.0, 3.0),
                                        (0.5, 5.0, 1.0, 3.0)):
        got = db.scalar("SELECT ROUND(? * ? / ?) * ?", (percent, one_rm, step,
                                                        step))
        mine = workouts.round_to(percent * one_rm, step)
        ties.append((percent, one_rm, step, got, mine))
        check(f"{percent * 100:g}% of {one_rm:g} at {step:g}: SQL {got}, "
              f"Python {mine}", (got, mine), (want, want))

    # --- a plan built by hand ---------------------------------------------
    print("\nThe write path")
    made = wm.save_plan({"name": "Built by hand", "rounding_kg": 5})
    week = wm.save_week(made["id"], {"number": 1})
    lookup = {row["name"]: row["id"] for row in wq.exercises()}
    wm.set_max(made["id"], lookup["Squats"], 140)
    session = wm.save_session(week["id"], {"number": 1}, [
        {"exercise_id": lookup["Squats"], "sets": [
            {"set_type": "warmup", "reps": "5", "load_mode": "percent",
             "percent_1rm": "50"},
            *[{"set_type": "working", "reps": "5", "load_mode": "percent",
               "percent_1rm": "80"} for _ in range(3)]]},
        {"exercise_id": lookup["Tricep Dips"], "sets": [
            {"set_type": "working", "reps": "8-10",
             "load_mode": "bodyweight", "added_kg": "7.5"}]},
        {"exercise_id": lookup["Walking Lunge (Dumbbells)"], "sets": [
            {"set_type": "accessory", "reps": "12", "load_mode": "choose"}]},
    ])
    # 1 warm-up + 3 working on squats, 1 bodyweight dip set, 1 accessory.
    check("saved", (session["exercises"], session["sets"]), (3, 6))
    sheet_rows = wq.session_sheet(session["id"])
    check("named after its working lifts",
          workouts.session_title(wq.session(session["id"])),
          "Session 1 - Squats + Tricep Dips")
    check("rounds to this plan's 5 kg, not the default 2.5",
          [row["prescribed_kg"] for row in sheet_rows[0]["sets"]],
          [70.0, 110.0, 110.0, 110.0])
    check("a per-side, per-dumbbell movement says so",
          (sheet_rows[2]["resolved_reps_mode"],
           sheet_rows[2]["resolved_weight_mode"]),
          ("per_side", "per_dumbbell"))
    check("and reads that way",
          workouts.fmt_reps(12, None, sheet_rows[2]["resolved_reps_mode"]),
          "12 each side")

    print("\nWhat the write path refuses")
    for label, call in [
        ("a percentage of a bodyweight movement", lambda: wm.save_session(
            week["id"], {"number": 2},
            [{"exercise_id": lookup["Pull-Ups"], "sets": [
                {"set_type": "working", "reps": "5", "load_mode": "percent",
                 "percent_1rm": "80"}]}])),
        ("bodyweight on a barbell movement", lambda: wm.save_session(
            week["id"], {"number": 2},
            [{"exercise_id": lookup["Squats"], "sets": [
                {"set_type": "working", "reps": "5",
                 "load_mode": "bodyweight"}]}])),
        ("four warm-up sets", lambda: wm.save_session(
            week["id"], {"number": 2},
            [{"exercise_id": lookup["Squats"], "sets": [
                {"set_type": "warmup", "reps": "5", "load_mode": "choose"}
                for _ in range(4)]}])),
        ("an exercise with no sets", lambda: wm.save_session(
            week["id"], {"number": 2},
            [{"exercise_id": lookup["Squats"], "sets": []}])),
        ("a session with no exercises", lambda: wm.save_session(
            week["id"], {"number": 2}, [])),
        ("a second plan of the same name",
         lambda: wm.save_plan({"name": "Built by hand"})),
        ("deleting an exercise a plan uses",
         lambda: wm.delete_exercise(lookup["Squats"])),
        ("a rep range that runs backwards", lambda: wm.save_session(
            week["id"], {"number": 2},
            [{"exercise_id": lookup["Squats"], "sets": [
                {"set_type": "working", "reps": "12-8",
                 "load_mode": "choose"}]}])),
    ]:
        try:
            call()
        except workouts.InvalidWorkout:
            check(f"refuses {label}", True, True)
        else:
            check(f"refuses {label}", False, True)

    # --- copying ----------------------------------------------------------
    print("\nCopying a plan takes the structure, not the history")
    wm.tick_session(session["id"], True)
    before = wq.totals(made["id"])
    copied = wm.copy_plan(made["id"], "Built by hand (copy)")
    after = wq.totals(copied["id"])
    check("same weeks", after["weeks"], before["weeks"])
    check("same sessions", after["sessions"], before["sessions"])
    check("same sets", after["sets"], before["sets"])
    check("nothing ticked off", after["sessions_done"], 0)
    check("the original still is", before["sessions_done"], 1)
    check("the 1RMs came too",
          [row["one_rm_kg"] for row in wq.maxes(copied["id"])], [140.0])

    print("\nCopying a week rebuilds the rotation")
    source = wq.weeks(plan["id"])[0]
    made_week = wm.copy_week(source["id"], plan["id"], 20)
    check("the copy holds the same sessions",
          len(wq.sessions(week_id=made_week["id"])),
          len(wq.sessions(week_id=source["id"])))
    check("and the same sets",
          db.scalar("SELECT COUNT(*) FROM v_exercise_sets WHERE week_id = ?",
                    (made_week["id"],), default=0),
          db.scalar("SELECT COUNT(*) FROM v_exercise_sets WHERE week_id = ?",
                    (source["id"],), default=0))

    print("\nThe tracker")
    imported = wq.totals(plan["id"])
    check("the workbook's ticks came over", result["sessions_ticked"], 20)
    check("which is weeks 1-10, both sessions",
          imported["sessions_done"] >= 20, True)
    nxt = wq.next_session(plan["id"])
    check("next up is the first one not done",
          (nxt["week_number"], nxt["number"]), (11, 1))

    print()
    if failures:
        print(f"FAILED: {len(failures)} check(s): {', '.join(failures)}")
        return 1
    print("The workbook reconciles: every set, every weight, and the two "
          "rounding rules agree.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
